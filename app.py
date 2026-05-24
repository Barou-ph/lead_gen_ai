"""
Lead Intelligence Tool v5 — Streamlit Edition
Deploy free trên Streamlit Cloud.

Chạy local: streamlit run app.py
"""

import os, json, re, random, time, threading, io
from datetime import datetime
import pandas as pd
import streamlit as st

# ══════════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Lead.AI — Lead Intelligence Tool",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════
#  CUSTOM CSS — giữ dark theme gần giống bản gốc
# ══════════════════════════════════════════════════════════════════════
st.markdown(
    """
<style>
[data-testid="stSidebar"] { background: #12151f; }
[data-testid="stSidebar"] .stMarkdown { color: #e2e8f0; }
.stTabs [data-baseweb="tab-list"] { background: #12151f; gap: 4px; }
.stTabs [data-baseweb="tab"] { background: #1e2235; color: #94a3b8; border-radius: 6px 6px 0 0; }
.stTabs [aria-selected="true"] { background: #161929; color: #fff; border-bottom: 2px solid #3b82f6; }
.grade-A { color: #eab308; font-weight: 700; }
.grade-B { color: #22c55e; font-weight: 700; }
.grade-C { color: #3b82f6; font-weight: 700; }
.grade-D { color: #ef4444; font-weight: 700; }
.metric-box { background: #161929; border: 1px solid #1e2235; border-radius: 8px;
              padding: 12px 16px; text-align: center; }
</style>
""",
    unsafe_allow_html=True,
)

# ══════════════════════════════════════════════════════════════════════
#  INDUSTRY META
# ══════════════════════════════════════════════════════════════════════
INDUSTRY_META = {
    "manufacturing": {
        "label": "🏭 Manufacturing",
        "desc": "Nhà máy, sản xuất, cơ khí",
        "why": "Ít noise nhất — yellowpages có danh sách nhà máy thật 100%",
    },
    "logistics": {
        "label": "🚚 Logistics",
        "desc": "Vận tải, kho bãi, freight",
        "why": "Công ty logistics VN nhiều, dễ tìm email sales",
    },
    "hospitality": {
        "label": "🏨 Hospitality",
        "desc": "Khách sạn, nhà hàng, du lịch",
        "why": "Team building là nhu cầu cốt lõi, budget tốt",
    },
    "healthcare": {
        "label": "🏥 Healthcare",
        "desc": "Phòng khám, dược phẩm, spa",
        "why": "Private clinic + pharma — ngân sách lớn, hay tổ chức team building",
    },
    "realestate": {
        "label": "🏗️ Real Estate",
        "desc": "BĐS, xây dựng, nội thất",
        "why": "Sales team lớn, incentive trip là nhu cầu thường xuyên",
    },
    "finance": {
        "label": "💰 Finance",
        "desc": "Tài chính, bảo hiểm, kế toán",
        "why": "Budget lớn, chuộng team building & retreat cao cấp",
    },
    "retail": {
        "label": "🛒 Retail",
        "desc": "Bán lẻ, chuỗi, siêu thị",
        "why": "Chuỗi nhiều nhân viên, cần gắn kết thường xuyên",
    },
    "education": {
        "label": "📚 Education",
        "desc": "Trường, trung tâm đào tạo",
        "why": "Trung tâm tư thục có ngân sách, cần team building cho staff",
    },
    "it": {
        "label": "💻 IT / Tech",
        "desc": "Phần mềm, outsourcing, agency",
        "why": "Công ty IT VN nhiều, developer burnout = nhu cầu team building cao",
    },
}

# ══════════════════════════════════════════════════════════════════════
#  YELLOWPAGES CATEGORIES
# ══════════════════════════════════════════════════════════════════════
YP_CATEGORIES = {
    "manufacturing": [
        "https://www.yellowpages.vn/cls/152060/co-khi----gia-cong-va-che-tao.html",
        "https://www.yellowpages.vn/cls/111010/do-go-noi-that---san-xuat-va-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/47910/hoa-chat---san-xuat,-nhap-khau-va-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/186060/nhua---cac-cong-ty-nhua.html",
        "https://www.yellowpages.vn/cls/174170/bao-bi---nha-san-xuat-va-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/225100/thep---cong-ty-thep-(san-xuat,-kinh-doanh).html",
        "https://www.yellowpages.vn/cls/64610/thuc-pham---san-xuat-va-che-bien.html",
        "https://www.yellowpages.vn/cls/97160/vai-soi---san-xuat-va-kinh-doanh.html",
    ],
    "hospitality": [
        "https://www.yellowpages.vn/cls/127160/khach-san.html",
        "https://www.yellowpages.vn/cls/200710/nha-hang.html",
        "https://www.yellowpages.vn/cls/51810/du-lich---cong-ty-lu-hanh-va-du-lich.html",
        "https://www.yellowpages.vn/cls/106010/thuc-pham---cung-cap-thuc-pham,-cong-ty-thuc-pham.html",
        "https://www.yellowpages.vn/cls/56010/cafe---quan-ca-phe-va-tra-sua.html",
    ],
    "logistics": [
        "https://www.yellowpages.vn/cls/246160/van-tai---cong-ty-van-tai-va-dai-ly-van-tai.html",
        "https://www.yellowpages.vn/cls/485215/van-tai-duong-bo.html",
        "https://www.yellowpages.vn/cls/213810/van-tai-bien.html",
        "https://www.yellowpages.vn/cls/68660/van-tai-container.html",
        "https://www.yellowpages.vn/cls/130610/giao-nhan-hang-hoa---chuyen-phat.html",
        "https://www.yellowpages.vn/cls/48310/kho-bai---cho-thue-va-quan-ly-kho.html",
        "https://www.yellowpages.vn/cls/135510/logistics---cong-ty-logistics.html",
    ],
    "healthcare": [
        "https://www.yellowpages.vn/cls/180660/duoc-pham---cong-ty-duoc-pham.html",
        "https://www.yellowpages.vn/cls/152660/y-te---benh-vien-va-co-so-chuyen-khoa.html",
        "https://www.yellowpages.vn/cls/424960/phong-kham-da-khoa.html",
        "https://www.yellowpages.vn/cls/77080/phong-kham-nha-khoa.html",
        "https://www.yellowpages.vn/cls/152560/thiet-bi-y-te---san-xuat,-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/126660/spa---tham-my-vien.html",
    ],
    "realestate": [
        "https://www.yellowpages.vn/cls/192550/bat-dong-san---cac-cong-ty-bat-dong-san.html",
        "https://www.yellowpages.vn/cls/37210/xay-dung---cong-ty-xay-dung.html",
        "https://www.yellowpages.vn/cls/420940/do-noi-that---thiet-ke-va-san-xuat.html",
        "https://www.yellowpages.vn/cls/197660/bat-dong-san---quan-ly-va-tu-van-bat-dong-san.html",
        "https://www.yellowpages.vn/cls/169260/van-phong---cho-thue-van-phong-(tron-goi,-ao,-co---working-space).html",
    ],
    "finance": [
        "https://www.yellowpages.vn/cls/131560/bao-hiem---cong-ty-bao-hiem.html",
        "https://www.yellowpages.vn/cls/10310/ke-toan-va-kiem-toan.html",
        "https://www.yellowpages.vn/cls/488245/ke-toan-thue---dich-vu-ke-toan-thue.html",
        "https://www.yellowpages.vn/cls/100310/tai-chinh---cac-cong-ty-tai-chinh.html",
        "https://www.yellowpages.vn/cls/38760/tu-van-quan-ly-doanh-nghiep.html",
    ],
    "retail": [
        "https://www.yellowpages.vn/cls/229210/sieu-thi.html",
        "https://www.yellowpages.vn/cls/225060/may-mac---quan-ao-thoi-trang.html",
        "https://www.yellowpages.vn/cls/116010/thuc-pham---san-xuat-va-phan-phoi.html",
        "https://www.yellowpages.vn/cls/147410/sieu-thi---trung-tam-thuong-mai.html",
    ],
    "education": [
        "https://www.yellowpages.vn/cls/245665/dao-tao---cac-cong-ty-dao-tao.html",
        "https://www.yellowpages.vn/cls/335515/dao-tao---ngoai-ngu.html",
        "https://www.yellowpages.vn/cls/487018/trung-tam-dao-tao.html",
        "https://www.yellowpages.vn/cls/140360/truong-mam-non---truong-mau-giao.html",
        "https://www.yellowpages.vn/cls/164810/ngoai-ngu---trung-tam-ngoai-ngu.html",
    ],
    "it": [
        "https://www.yellowpages.vn/cls/66030/phan-mem.html",
        "https://www.yellowpages.vn/cls/226260/cong-nghe-thong-tin---it-services.html",
        "https://www.yellowpages.vn/cls/166210/thiet-ke-web---cong-ty-thiet-ke-website.html",
        "https://www.yellowpages.vn/cls/231510/digital-marketing---agency.html",
        "https://www.yellowpages.vn/cls/373705/phan-mem-quan-tri.html",
        "https://www.yellowpages.vn/cls/493201/phan-mem---dich-vu-gia-cong-phan-mem.html",
    ],
}

SKIP_DOMAINS_YP = {
    "yellowpages.vn",
    "yellowpages.com.vn",
    "google.com",
    "facebook.com",
    "zalo.me",
    "youtube.com",
    "twitter.com",
    "instagram.com",
    "tiktok.com",
    "linkedin.com",
    "apple.com",
    "microsoft.com",
    "maps.google.com",
    "trangvangvietnam.com",
    "signup.trangvangvietnam.com",
    "tiki.vn",
    "shopee.vn",
    "lazada.vn",
    "sendo.vn",
    "thegioididong.com",
    "fptshop.com.vn",
    "dienmayxanh.com",
    "vnexpress.net",
    "tuoitre.vn",
    "thanhnien.vn",
    "dantri.com.vn",
    "wikipedia.org",
    "topcv.vn",
    "itviec.com",
    "vietnamworks.com",
    "vieclam24h.vn",
}

# ══════════════════════════════════════════════════════════════════════
#  FILTER CONSTANTS
# ══════════════════════════════════════════════════════════════════════
PERSONAL_DOMAINS = {
    "gmail.com",
    "yahoo.com",
    "hotmail.com",
    "outlook.com",
    "icloud.com",
    "mail.com",
    "protonmail.com",
    "yandex.com",
    "live.com",
    "msn.com",
    "aol.com",
}

NOISE_DOMAINS = {
    "vnexpress.net",
    "tuoitre.vn",
    "thanhnien.vn",
    "dantri.com.vn",
    "vietcetera.com",
    "techcrunch.com",
    "forbes.com",
    "bloomberg.com",
    "reuters.com",
    "linkedin.com",
    "itviec.com",
    "vietnamworks.com",
    "topcv.vn",
    "jobstreet.com",
    "gov.vn",
    "chinhphu.vn",
    "shopee.vn",
    "lazada.vn",
    "tiki.vn",
    "thegioididong.com",
    "medium.com",
    "substack.com",
    "wikipedia.org",
    "trangvangvietnam.com",
    "yellowpages.com.vn",
    "yellowpages.vn",
    "tripadvisor.com",
    "booking.com",
    "agoda.com",
    "reddit.com",
    "quora.com",
}

REJECT_DESC_PHRASES = [
    "latest news",
    "breaking news",
    "tin tức mới",
    "read more articles",
    "market size",
    "market share",
    "cagr",
    "table of contents",
    "buy report",
    "download report",
    "find companies",
    "list of companies",
    "top companies in",
    "best companies in",
    "danh sách công ty",
    "online games",
    "play free",
    "scholarship",
    "tuition",
    "admissions",
    "tin tức",
    "báo điện tử",
    "toà soạn",
    "giấy phép mạng xã hội",
]

BAD_EMAIL_PATTERNS = [
    r"sentry",
    r"wixpress",
    r"bug-report",
    r"^name@",
    r"^your@",
    r"^youname@",
    r"@email\.com$",
    r"@yourcompany\.com$",
    r"@example\.com$",
    r"^test@",
    r"^demo@",
    r"@company\.com$",
    r"@mail\.com$",
    r"^[a-f0-9]{20,}@",
    r"@robots\.net$",
    r"@avif$",
    r"^toasoan@",
    r"^ads@",
    r"@sohu\.com$",
    r"@qq\.com$",
    r"@163\.com$",
    r"@126\.com$",
    r"@sina\.com$",
    r"^\d{5,}@",
    r"^user@",
    r"^example@",
    r"@myorg\.com$",
]

BAD_EMAIL_DOMAINS_EXTRA = {
    "sentry.io",
    "wixpress.com",
    "example.com",
    "yourcompany.com",
    "robots.net",
    "schema.org",
    "w3.org",
    "cloudflare.com",
    "sohu.com",
    "qq.com",
    "163.com",
    "126.com",
    "sina.com",
    "myorg.com",
    "trangvangvietnam.com",
}

FOREIGN_EMAIL_TLDS = {".fi", ".bd", ".lk", ".pk", ".th", ".cn"}


# ══════════════════════════════════════════════════════════════════════
#  FILTER HELPERS
# ══════════════════════════════════════════════════════════════════════
def _get_domain(url):
    try:
        from urllib.parse import urlparse

        return urlparse(str(url)).netloc.lower().replace("www.", "")
    except:
        return ""


def _root_domain(url):
    try:
        parts = str(url).split("//")[-1].split("/")[0].split(".")
        return ".".join(parts[-2:]) if len(parts) >= 2 else str(url)
    except:
        return str(url)


def _is_noise_domain(url):
    if not url:
        return True
    dom = _get_domain(url)
    if dom in NOISE_DOMAINS:
        return True
    for nd in NOISE_DOMAINS:
        if dom.endswith("." + nd):
            return True
    return False


def _desc_ok(desc):
    if not desc or str(desc).strip() in ("", "nan", "None"):
        return True
    low = desc.lower()
    return not any(p in low for p in REJECT_DESC_PHRASES)


def _is_bad_email(email):
    if not email or "@" not in email or len(email) > 100:
        return True
    low = email.lower()
    domain = low.split("@")[-1]
    prefix = low.split("@")[0]
    if domain in BAD_EMAIL_DOMAINS_EXTRA:
        return True
    for tld in FOREIGN_EMAIL_TLDS:
        if domain.endswith(tld):
            return True
    if any(re.search(p, low) for p in BAD_EMAIL_PATTERNS):
        return True
    if len(prefix) > 50:
        return True
    return False


def _email_type(email):
    if not email or "@" not in email:
        return "none"
    dom = email.split("@")[-1].lower().strip()
    return "personal" if dom in PERSONAL_DOMAINS else "company"


def _has_value(val):
    return bool(
        val and str(val).strip() not in ("", "nan", "None", "NaN", "false", "False")
    )


def _reject_reason(d, cfg):
    if _is_noise_domain(str(d.get("website", ""))):
        return "noise_domain"
    if not _desc_ok(str(d.get("description", ""))):
        return "bad_description"
    grade = str(d.get("grade", "")).strip().upper()
    if grade not in cfg["grades"]:
        return f"grade_{grade or 'missing'}"
    try:
        score = float(d.get("score", 0) or 0)
    except:
        score = 0
    if score < cfg["min_score"]:
        return f"score_{int(score)}"
    website = str(d.get("website", "")).strip()
    if not website or not website.startswith("http"):
        return "no_website"
    if cfg["require_email"]:
        em = str(d.get("best_email", "") or d.get("emails", "") or "").strip()
        first = em.split(",")[0].strip() if em else ""
        if not first or _is_bad_email(first):
            return "bad_email"
    if cfg["require_phone"]:
        if not _has_value(d.get("phones", "")):
            return "no_phone"
    if cfg["industries"]:
        ind = str(d.get("industry", "")).lower()
        if ind not in [i.lower() for i in cfg["industries"]]:
            return f"industry_{ind or '?'}"
    return ""


def apply_filter(df, cfg):
    passed, rejected = [], []
    seen_emails = set()
    seen_domains = set()
    for _, row in df.iterrows():
        d = row.to_dict()
        reason = _reject_reason(d, cfg)
        if reason:
            rejected.append({**d, "reject_reason": reason})
            continue
        em = (
            str(d.get("best_email", "") or d.get("emails", "") or "")
            .split(",")[0]
            .strip()
            .lower()
        )
        dom = _root_domain(str(d.get("website", "")))
        if em and em not in ("", "nan", "none"):
            if em in seen_emails:
                rejected.append({**d, "reject_reason": "duplicate_email"})
                continue
            seen_emails.add(em)
        if dom:
            if dom in seen_domains:
                rejected.append({**d, "reject_reason": "duplicate_domain"})
                continue
            seen_domains.add(dom)
        passed.append(d)

    df_pass = pd.DataFrame(passed) if passed else pd.DataFrame()
    df_rej = pd.DataFrame(rejected) if rejected else pd.DataFrame()

    if not df_pass.empty:

        def _fe(r):
            em = str(r.get("best_email", "") or r.get("emails", "") or "")
            return em.split(",")[0].strip() if em else ""

        df_pass["email_type"] = df_pass.apply(
            lambda r: _email_type(_fe(r.to_dict())), axis=1
        )
        go = {"A": 0, "B": 1, "C": 2, "D": 3}
        df_pass["_g"] = df_pass.get("grade", pd.Series(["D"] * len(df_pass))).apply(
            lambda g: go.get(str(g).upper(), 3)
        )
        df_pass["_s"] = pd.to_numeric(df_pass.get("score", 0), errors="coerce").fillna(
            0
        )
        df_pass.sort_values(["_g", "_s"], ascending=[True, False], inplace=True)
        df_pass.drop(columns=["_g", "_s"], inplace=True, errors="ignore")
        df_pass.reset_index(drop=True, inplace=True)

    return df_pass, df_rej


# ══════════════════════════════════════════════════════════════════════
#  CRAWLER (chạy trong thread, cập nhật session_state)
# ══════════════════════════════════════════════════════════════════════
def _get_dom_simple(url):
    try:
        return url.split("//")[-1].split("/")[0].lower().replace("www.", "")
    except:
        return ""


def run_crawl_sync(industry, target, workers, skip_enrich):
    """Hàm crawl đồng bộ, trả về list leads. Chạy trong thread riêng."""
    import requests, urllib3
    from concurrent.futures import ThreadPoolExecutor, as_completed

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return [], ["❌ Thiếu thư viện: pip install beautifulsoup4 requests"]

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    logs = []
    session_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    UA = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    ]

    def hdrs():
        return {
            "User-Agent": random.choice(UA),
            "Accept": "text/html,*/*;q=0.8",
            "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.8",
            "Referer": "https://www.yellowpages.vn/",
        }

    def fetch(url, timeout=12):
        for attempt in range(2):
            try:
                r = requests.get(
                    url,
                    headers=hdrs(),
                    timeout=timeout,
                    verify=False,
                    allow_redirects=True,
                )
                if r.status_code == 200:
                    return r
                if r.status_code in (429, 503):
                    time.sleep(5 * (attempt + 1))
            except:
                if attempt == 0:
                    time.sleep(2)
        return None

    def extract_emails(text):
        raw = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
        return list(dict.fromkeys(e.lower() for e in raw))

    def extract_phones(text):
        raw = re.findall(r"(?:0|\+84)[0-9\s\.\-]{8,12}[0-9]", text)
        cleaned = []
        for p in raw:
            p2 = re.sub(r"[\s\.\-]", "", p)
            if 9 <= len(p2) <= 12:
                cleaned.append(p2)
        return list(dict.fromkeys(cleaned))

    def meta_desc(html):
        try:
            soup = BeautifulSoup(html, "html.parser")
            for attr in [{"name": "description"}, {"property": "og:description"}]:
                tag = soup.find("meta", attrs=attr)
                if tag and tag.get("content"):
                    return tag["content"][:300].strip()
            for p in soup.find_all("p"):
                t = p.get_text(strip=True)
                if len(t) > 60:
                    return t[:300]
        except:
            pass
        return ""

    # Phase 1: lấy /lgs/ links
    logs.append(f"📋 PHASE 1: Lấy danh sách từ YellowPages ({industry})")
    cats = YP_CATEGORIES.get(industry, [])
    lgs_links = []
    seen_lgs = set()

    for cat_url in cats:
        # Check nếu bị stop
        if st.session_state.get("crawl_stop", False):
            break
        slug = cat_url.split("/")[-1][:50]
        logs.append(f"  📂 {slug}")
        for page in range(1, 6):
            if st.session_state.get("crawl_stop", False):
                break
            url = cat_url if page == 1 else cat_url.replace(".html", f"-p{page}.html")
            r = fetch(url)
            if not r:
                break
            soup = BeautifulSoup(r.text, "html.parser")
            found = 0
            for a in soup.select("a[href]"):
                href = a.get("href", "")
                if "/lgs/" in href:
                    full = (
                        "https://www.yellowpages.vn" + href
                        if href.startswith("/")
                        else href
                    )
                    if full not in seen_lgs:
                        seen_lgs.add(full)
                        lgs_links.append((full, industry))
                        found += 1
            logs.append(f"    trang {page}: +{found} ({len(lgs_links)} total)")
            if found == 0:
                break
            time.sleep(random.uniform(0.8, 1.5))
        if len(lgs_links) >= target * 3:
            break

    logs.append(f"✅ {len(lgs_links)} công ty trong danh sách\n")
    st.session_state["crawl_logs"] = logs[:]

    if not lgs_links:
        return [], logs

    # Phase 2: lấy website từ mỗi trang /lgs/
    logs.append("🔗 PHASE 2: Lấy website thật từ trang công ty")
    website_queue = []
    seen_domains = set()

    for i, (lgs_url, ind) in enumerate(lgs_links):
        if st.session_state.get("crawl_stop", False):
            break
        r = fetch(lgs_url)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        company_name = ""
        h1 = soup.find("h1")
        if h1:
            company_name = h1.get_text(strip=True)
        website = ""
        for a in soup.select("a[href]"):
            href = a.get("href", "")
            if not href.startswith("http"):
                continue
            dom = _get_dom_simple(href)
            if dom and dom not in SKIP_DOMAINS_YP and len(dom) > 4:
                website = href.split("?")[0].rstrip("/")
                break
        if website:
            dom = _get_dom_simple(website)
            if dom and dom not in seen_domains and dom not in SKIP_DOMAINS_YP:
                seen_domains.add(dom)
                website_queue.append((website, ind, company_name))
                logs.append(f"  ✅ [{i+1}] {company_name[:30]} → {website[:50]}")
        time.sleep(random.uniform(0.4, 0.8))
        if len(website_queue) >= target * 2:
            break

    st.session_state["crawl_logs"] = logs[:]
    logs.append(f"\n✅ {len(website_queue)} websites để crawl\n")

    if not website_queue:
        return [], logs

    # Phase 3: crawl từng website
    logs.append(f"🕷️  PHASE 3: CRAWL {len(website_queue)} websites ({workers} threads)")
    st.session_state["crawl_total"] = len(website_queue)
    raw_leads = []
    done_count = [0]

    def crawl_one(item):
        ws, ind, name = item
        try:
            dom = _get_dom_simple(ws)
            if not dom or dom in SKIP_DOMAINS_YP:
                return None
            r = fetch(ws, timeout=12)
            if not r:
                return None
            html = r.text
            desc = meta_desc(html)
            try:
                desc = desc.encode("latin-1").decode("utf-8")
            except:
                pass
            if desc.count("á»") > 2:
                desc = ""
            emails_raw = extract_emails(html)
            phones_raw = extract_phones(html)
            emails = [e for e in emails_raw if not _is_bad_email(e)][:5]
            phones = phones_raw[:3]
            if not emails and not phones:
                return None
            is_vn = dom.endswith(".vn") or ".vn." in dom
            if not is_vn and not phones:
                return None
            best_email = emails[0] if emails else ""
            return {
                "website": ws,
                "company_name": name,
                "best_email": best_email,
                "emails": ", ".join(emails),
                "phones": ", ".join(phones),
                "description": desc,
                "industry": ind,
                "grade": "B",
                "score": 30 + (10 if best_email else 0) + (5 if phones else 0),
                "field": ind,
                "session": session_id,
            }
        except:
            return None

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(crawl_one, item): item for item in website_queue}
        for fut in as_completed(futures):
            done_count[0] += 1
            st.session_state["crawl_done"] = done_count[0]
            result = None
            try:
                result = fut.result()
            except:
                pass
            bar = f"[{done_count[0]}/{len(website_queue)}|{len(raw_leads)}↑]"
            if result:
                raw_leads.append(result)
                logs.append(f"  ✅ {bar} {result['website'][:50]}")
            else:
                item = futures[fut]
                logs.append(f"  ❌ {bar} {item[0][:55]}")

    logs.append(f"\n🎉 DONE | {len(raw_leads)} leads | {industry.upper()}")
    return raw_leads, logs


# ══════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════
def df_to_excel_bytes(df_pass, df_rej=None):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HFILL = PatternFill("solid", fgColor="1F4E79")
    HFONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    DFONT = Font(name="Calibri", size=9)
    THIN = Side(style="thin", color="CCCCCC")
    BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR_A = Alignment(horizontal="center", vertical="center")
    LFT_A = Alignment(horizontal="left", vertical="top", wrap_text=True)
    GRADE_COLORS = {"A": "FFF2CC", "B": "E2EFDA", "C": "DDEEFF", "D": "FFE0E0"}

    wb = Workbook()

    def write_sheet(ws, df, extra_cols=None):
        if df is None or df.empty:
            ws["A1"] = "Không có dữ liệu."
            return
        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(1, ci, col)
            cell.font, cell.fill, cell.alignment, cell.border = HFONT, HFILL, CTR_A, BDR
            ws.column_dimensions[get_column_letter(ci)].width = max(
                12, min(len(col) * 2, 50)
            )
        for ri, row in enumerate(df.itertuples(index=False), 2):
            grade = str(
                getattr(row, "grade", "") if "grade" in df.columns else ""
            ).upper()
            rfill = PatternFill("solid", fgColor=GRADE_COLORS.get(grade, "FFFFFF"))
            for ci, col in enumerate(cols, 1):
                val = getattr(row, col, "") if col in df.columns else ""
                if pd.isna(val) if not isinstance(val, str) else False:
                    val = ""
                cell = ws.cell(ri, ci, str(val) if val != "" else "")
                cell.font, cell.fill, cell.border = DFONT, rfill, BDR
                cell.alignment = (
                    LFT_A if col in {"description", "email_body", "emails"} else CTR_A
                )
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "✅ Leads"
    write_sheet(ws1, df_pass)

    if df_rej is not None and not df_rej.empty:
        ws2 = wb.create_sheet("❌ Rejected")
        write_sheet(ws2, df_rej)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════
#  SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════
for key, val in {
    "crawl_leads": [],
    "crawl_logs": [],
    "crawl_running": False,
    "crawl_done": 0,
    "crawl_total": 0,
    "crawl_stop": False,
    "filter_df": None,
    "filter_passed": None,
    "filter_rejected": None,
    "active_tab": "crawl",
}.items():
    if key not in st.session_state:
        st.session_state[key] = val

# ══════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🎯 Lead.AI `v5 · YP`")
    st.markdown("---")

    active_tab = st.radio(
        "Điều hướng",
        ["🕷️ Crawl", "🔍 Filter", "📧 Email", "❌ Rejected"],
        label_visibility="collapsed",
    )
    st.session_state["active_tab"] = active_tab
    st.markdown("---")

    # ── Crawl config ──────────────────────────────────────────────────
    if "Crawl" in active_tab:
        st.markdown("### Chọn ngành")
        industry_keys = list(INDUSTRY_META.keys())
        industry_labels = [
            f"{INDUSTRY_META[k]['label']} ({len(YP_CATEGORIES.get(k,[]))} cat)"
            for k in industry_keys
        ]
        idx = st.selectbox(
            "Ngành",
            range(len(industry_keys)),
            format_func=lambda i: industry_labels[i],
            index=2,
            label_visibility="collapsed",
        )
        selected_industry = industry_keys[idx]
        meta = INDUSTRY_META[selected_industry]
        st.info(f"**{meta['desc']}**\n\n💡 {meta['why']}")

        st.markdown("### Cấu hình")
        col1, col2 = st.columns(2)
        target = col1.number_input("Target leads", 10, 500, 60, step=10)
        workers = col2.number_input("Threads", 1, 12, 6)

        skip_enrich = not st.checkbox("🔍 Enrich pages", value=True)

        if not st.session_state["crawl_running"]:
            if st.button("▶ Bắt đầu crawl", type="primary", use_container_width=True):
                st.session_state["crawl_stop"] = False
                st.session_state["crawl_running"] = True
                st.session_state["crawl_logs"] = []
                st.session_state["crawl_leads"] = []
                st.session_state["crawl_done"] = 0
                st.session_state["crawl_total"] = 0

                def _bg(ind, tgt, wrk, skip):
                    leads, logs = run_crawl_sync(ind, tgt, wrk, skip)
                    st.session_state["crawl_leads"] = leads
                    st.session_state["crawl_logs"] = logs
                    st.session_state["crawl_running"] = False
                    # Auto load vào filter
                    if leads:
                        st.session_state["filter_df"] = pd.DataFrame(leads)

                t = threading.Thread(
                    target=_bg,
                    args=(selected_industry, int(target), int(workers), skip_enrich),
                    daemon=True,
                )
                t.start()
                st.rerun()
        else:
            if st.button("⏹ Dừng crawl", type="secondary", use_container_width=True):
                st.session_state["crawl_stop"] = True

    # ── Filter config ─────────────────────────────────────────────────
    elif "Filter" in active_tab:
        st.markdown("### Nguồn dữ liệu")
        uploaded = st.file_uploader(
            "Upload Excel / CSV",
            type=["xlsx", "xls", "csv"],
            label_visibility="collapsed",
        )
        if uploaded:
            try:
                df = (
                    pd.read_csv(uploaded, dtype=str).fillna("")
                    if uploaded.name.endswith(".csv")
                    else pd.read_excel(uploaded, dtype=str).fillna("")
                )
                st.session_state["filter_df"] = df
                st.success(f"✅ {uploaded.name} — {len(df)} rows")
            except Exception as e:
                st.error(str(e))

        if st.session_state["crawl_leads"] and st.session_state["filter_df"] is None:
            st.info(f"Dùng {len(st.session_state['crawl_leads'])} leads từ Crawl")

        st.markdown("### Filter Config")
        min_score = st.slider("Min Score", 0, 100, 15)
        grades = st.multiselect("Grades", ["A", "B", "C", "D"], default=["A", "B", "C"])
        require_email = st.checkbox("Cần Email", value=True)
        require_phone = st.checkbox("Cần Phone", value=False)
        ind_filter = st.text_input(
            "Industry (hospitality, it...)", placeholder="Bỏ trống = tất cả"
        )

        if st.button("🔍 Lọc Leads", type="primary", use_container_width=True):
            df_src = st.session_state.get("filter_df")
            if df_src is None and st.session_state["crawl_leads"]:
                df_src = pd.DataFrame(st.session_state["crawl_leads"])
            if df_src is None:
                st.error("Chưa có dữ liệu! Upload file hoặc crawl trước.")
            else:
                cfg = {
                    "min_score": min_score,
                    "grades": [g.upper() for g in grades],
                    "require_email": require_email,
                    "require_phone": require_phone,
                    "industries": [
                        x.strip() for x in ind_filter.split(",") if x.strip()
                    ],
                }
                df_p, df_r = apply_filter(df_src.copy(), cfg)
                st.session_state["filter_passed"] = df_p
                st.session_state["filter_rejected"] = df_r
                st.success(f"✅ {len(df_p)} leads | ❌ {len(df_r)} rejected")
                st.rerun()

    # ── Email config ──────────────────────────────────────────────────
    elif "Email" in active_tab:
        st.markdown("### Gen Cold Email")
        n = len(st.session_state.get("filter_passed") or [])
        if n:
            st.success(f"{n} leads sẵn sàng")
            calls = max(1, n // 5 + 1)
            cost = (calls * 500 * 0.15 + calls * 1000 * 0.60) / 1_000_000
            st.caption(f"~${cost:.4f} USD ước tính ({calls} API calls)")
            if st.button("🤖 Gen AI Email", type="primary", use_container_width=True):
                st.info(
                    "Tính năng AI Email cần module `email_generator.py` — xem hướng dẫn tích hợp bên dưới."
                )
        else:
            st.warning("Filter leads trước để gen email.")
        st.markdown("""
**Hướng dẫn:**
1. Crawl → Filter leads
2. Bấm Gen AI Email
3. Export Excel → gửi
""")

# ══════════════════════════════════════════════════════════════════════
#  MAIN CONTENT
# ══════════════════════════════════════════════════════════════════════

# ── TAB: CRAWL ────────────────────────────────────────────────────────
if "Crawl" in active_tab:
    st.markdown("## 🕷️ Crawl YellowPages")

    # Progress bar
    if st.session_state["crawl_running"]:
        total = st.session_state["crawl_total"] or 1
        done = st.session_state["crawl_done"]
        pct = done / total
        st.progress(
            pct,
            text=f"⏳ Đang crawl... {done}/{total} URLs | {len(st.session_state['crawl_leads'])} leads",
        )
        st.button("🔄 Refresh", on_click=st.rerun)

    # Stats nhanh nếu có leads
    leads = st.session_state["crawl_leads"]
    if leads:
        gc = {"A": 0, "B": 0, "C": 0, "D": 0}
        for l in leads:
            g = l.get("grade", "D")
            gc[g] = gc.get(g, 0) + 1
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total", len(leads))
        c2.metric("Grade A", gc["A"])
        c3.metric("Grade B", gc["B"])
        c4.metric("Grade C", gc["C"])

        # Export button
        excel_bytes = df_to_excel_bytes(pd.DataFrame(leads))
        st.download_button(
            "⬇️ Export Excel",
            data=excel_bytes,
            file_name=f"leads_crawl_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Bảng leads
        df_show = pd.DataFrame(leads)[
            [
                c
                for c in [
                    "grade",
                    "score",
                    "website",
                    "best_email",
                    "emails",
                    "phones",
                    "industry",
                    "session",
                ]
                if c in pd.DataFrame(leads).columns
            ]
        ]
        st.dataframe(df_show, use_container_width=True, height=400)

    # Logs
    if st.session_state["crawl_logs"]:
        with st.expander("📋 Logs crawl", expanded=not leads):
            log_text = "\n".join(st.session_state["crawl_logs"][-200:])
            st.code(log_text, language=None)
    else:
        st.info("Chọn ngành bên trái và nhấn **▶ Bắt đầu crawl** để bắt đầu.")

# ── TAB: FILTER ───────────────────────────────────────────────────────
elif "Filter" in active_tab:
    st.markdown("## 🔍 Filter Leads")

    df_pass = st.session_state.get("filter_passed")
    df_rej = st.session_state.get("filter_rejected")

    if df_pass is not None and not df_pass.empty:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("✅ Passed", len(df_pass))
        c2.metric("❌ Rejected", len(df_rej) if df_rej is not None else 0)
        gc = (
            df_pass["grade"].value_counts().to_dict()
            if "grade" in df_pass.columns
            else {}
        )
        c3.metric("Grade A/B", gc.get("A", 0) + gc.get("B", 0))
        has_email = (
            df_pass["best_email"]
            .apply(lambda x: bool(str(x).strip() and str(x) not in ("", "nan", "None")))
            .sum()
            if "best_email" in df_pass.columns
            else 0
        )
        c4.metric("Có Email", has_email)

        # Export
        excel_bytes = df_to_excel_bytes(df_pass, df_rej)
        st.download_button(
            "⬇️ Export Excel (Passed + Rejected)",
            data=excel_bytes,
            file_name=f"leads_filtered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.markdown("### ✅ Leads đã lọc")
        st.dataframe(df_pass, use_container_width=True, height=500)
    else:
        df_src = st.session_state.get("filter_df")
        if df_src is not None:
            st.info(f"Có {len(df_src)} rows — nhấn **🔍 Lọc Leads** ở sidebar để lọc.")
            st.dataframe(df_src.head(20), use_container_width=True)
        elif st.session_state["crawl_leads"]:
            st.info(
                f"Có {len(st.session_state['crawl_leads'])} leads từ crawl — nhấn **🔍 Lọc Leads** ở sidebar."
            )
        else:
            st.info("Upload file Excel/CSV hoặc crawl trước, sau đó lọc ở sidebar.")

# ── TAB: EMAIL ────────────────────────────────────────────────────────
elif "Email" in active_tab:
    st.markdown("## 📧 Cold Email")

    df_pass = st.session_state.get("filter_passed")
    if df_pass is not None and not df_pass.empty:
        if "email_body" in df_pass.columns:
            with_email = df_pass[
                df_pass["email_body"].apply(
                    lambda x: bool(str(x).strip() and str(x) not in ("", "nan", "None"))
                )
            ]
            st.markdown(f"**{len(with_email)}/{len(df_pass)} leads đã có email**")
            for _, row in with_email.iterrows():
                with st.container(border=True):
                    try:
                        site = row.get("website", "")
                        from urllib.parse import urlparse

                        site = urlparse(site).netloc.replace("www.", "")
                    except:
                        site = row.get("website", "")
                    st.markdown(f"**🌐 {site}** · `{row.get('best_email','')}`")
                    st.markdown(f"**📧 {row.get('email_subject','(no subject)')}**")
                    st.code(str(row.get("email_body", "")), language=None)
                    if row.get("pain_point"):
                        st.caption(f"💡 {row.get('pain_point','')}")
        else:
            st.info(
                "Leads chưa có email. Dùng module `email_generator.py` để gen AI email."
            )

        # Export
        excel_bytes = df_to_excel_bytes(
            df_pass, st.session_state.get("filter_rejected")
        )
        st.download_button(
            "⬇️ Export Excel",
            data=excel_bytes,
            file_name=f"leads_email_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("Filter leads ở tab 🔍 Filter trước.")

# ── TAB: REJECTED ─────────────────────────────────────────────────────
elif "Rejected" in active_tab:
    st.markdown("## ❌ Rejected Leads")

    df_rej = st.session_state.get("filter_rejected")
    if df_rej is not None and not df_rej.empty:
        # Thống kê lý do
        if "reject_reason" in df_rej.columns:
            rc = df_rej["reject_reason"].value_counts().head(10)
            st.markdown("### Lý do bị loại")
            cols = st.columns(min(5, len(rc)))
            for i, (reason, cnt) in enumerate(rc.items()):
                cols[i % 5].metric(reason, cnt)

        st.markdown(f"### {len(df_rej)} leads bị loại")
        show_cols = [
            c
            for c in ["grade", "score", "website", "reject_reason", "field", "industry"]
            if c in df_rej.columns
        ]
        st.dataframe(
            df_rej[show_cols] if show_cols else df_rej,
            use_container_width=True,
            height=500,
        )
    else:
        st.info("Chưa có dữ liệu. Chạy Filter trước.")
