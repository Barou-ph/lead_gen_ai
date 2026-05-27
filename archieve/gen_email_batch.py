"""
gen_email_batch.py — Batch gen cold email cho top 100 leads
Chạy: python gen_email_batch.py

Workflow:
1. Đọc top100_leads.xlsx (chạy filter_top100.py trước)
2. Render cold email theo ngành (dùng email_templates.py)
3. Tùy chọn: gọi Anthropic API để personalize thêm (10 leads đầu)
4. Xuất top100_with_emails.xlsx
"""

import os, sys, json, time
import pandas as pd

# ── CONFIG ────────────────────────────────────────────────────────────
INPUT_FILE = "top100_leads.xlsx"  # Output của filter_top100.py
OUTPUT_FILE = "top100_with_emails.xlsx"
AI_BATCH = 10  # Số leads dùng AI personalize (tiết kiệm cost)
USE_AI = True  # Bật/tắt AI personalization

# ── ANTHROPIC CONFIG (nếu USE_AI=True) ───────────────────────────────
# Đặt API key vào .env hoặc biến môi trường ANTHROPIC_API_KEY
ANTHROPIC_MODEL = "claude-haiku-4-5-20251001"  # Rẻ nhất, đủ nhanh


# ── LOAD TEMPLATES ────────────────────────────────────────────────────
def load_templates():
    try:
        from email_templates import render_all, render_email, SENDER

        return render_all, render_email, SENDER
    except ImportError:
        print("❌ Không tìm thấy email_templates.py. Đặt file vào cùng thư mục.")
        sys.exit(1)


# ── AI PERSONALIZE ────────────────────────────────────────────────────
def ai_personalize_batch(leads: list, n: int = 10) -> list:
    """
    Gửi n leads đầu cho Claude để personalize subject line + mở đầu email.
    Giữ nguyên body template, chỉ cải thiện phần mở đầu.
    """
    try:
        import anthropic
    except ImportError:
        print("⚠️  anthropic package chưa cài. Chạy: pip install anthropic")
        return leads

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        try:
            from dotenv import load_dotenv

            load_dotenv()
            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        except:
            pass
    if not api_key:
        print("⚠️  Không có ANTHROPIC_API_KEY. Bỏ qua AI personalization.")
        return leads

    client = anthropic.Anthropic(api_key=api_key)
    batch = leads[:n]
    rest = leads[n:]
    result = []

    print(f"\n🤖 AI Personalize {n} leads đầu...")
    for i, lead in enumerate(batch):
        try:
            prompt = f"""Bạn là copywriter chuyên cold email B2B tiếng Việt cho dịch vụ team building.

Thông tin lead:
- Website: {lead.get('website', '')}
- Industry: {lead.get('industry', '')}
- Field: {lead.get('field', '')}
- Description: {lead.get('description', '')[:200]}
- Pain point: {lead.get('pain_point', '')}

Email hiện tại:
Subject: {lead.get('email_subject', '')}
Body (50 từ đầu): {lead.get('email_body', '')[:200]}

Nhiệm vụ:
1. Viết 1 subject line tốt hơn (max 60 ký tự, cụ thể, có tên company nếu biết)
2. Viết lại 2-3 câu mở đầu email (thay đoạn đầu hiện tại) — cụ thể hơn dựa trên description

Trả về JSON:
{{"subject": "...", "opening": "..."}}

Chỉ trả JSON, không giải thích."""

            msg = client.messages.create(
                model=ANTHROPIC_MODEL,
                max_tokens=300,
                messages=[{"role": "user", "content": prompt}],
            )
            text = msg.content[0].text.strip()
            # Parse JSON
            if "```" in text:
                text = text.split("```")[1].replace("json", "").strip()
            data = json.loads(text)
            # Patch email
            if data.get("subject"):
                lead["email_subject"] = data["subject"]
            if data.get("opening") and lead.get("email_body"):
                lines = lead["email_body"].split("\n")
                # Thay 2 dòng đầu (sau "Chào X,")
                greeting = lines[0] if lines else "Chào Anh/Chị,"
                rest_body = (
                    "\n".join(lines[3:]) if len(lines) > 3 else lead["email_body"]
                )
                lead["email_body"] = f"{greeting}\n\n{data['opening']}\n\n{rest_body}"
            lead["ai_personalized"] = True
            print(f"  ✅ [{i+1}/{n}] AI personalized: {lead.get('website','')[:40]}")
            time.sleep(0.5)  # Rate limit
        except Exception as e:
            print(f"  ⚠️  [{i+1}/{n}] AI skip: {e}")
            lead["ai_personalized"] = False
        result.append(lead)

    return result + rest


# ── EXPORT EXCEL ──────────────────────────────────────────────────────
def export_with_emails(df: pd.DataFrame, output_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        GRADE_CLR = {"A": "FFD700", "B": "C6EFCE", "C": "DDEEFF"}
        THIN = Side(style="thin", color="CCCCCC")
        BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        HFILL = PatternFill("solid", fgColor="1F4E79")
        HFONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        DFONT = Font(name="Calibri", size=9)
        CTR_A = Alignment(horizontal="center", vertical="center")
        LFT_A = Alignment(horizontal="left", vertical="center")
        WRAP_A = Alignment(horizontal="left", vertical="top", wrap_text=True)

        COLS = [
            ("grade", 7),
            ("score", 7),
            ("email_subject", 48),
            ("best_email_clean", 34),
            ("email_type", 12),
            ("website", 38),
            ("industry", 14),
            ("field", 18),
            ("email_body", 70),
            ("pain_point", 35),
            ("phones", 20),
            ("ai_personalized", 12),
            ("session", 18),
        ]
        WRAP_COLS = {"email_body", "pain_point", "description"}
        CTR_COLS = {"grade", "score", "email_type", "industry", "ai_personalized"}

        wb = Workbook()
        ws = wb.active
        ws.title = "Cold Emails"

        for ci, (col, width) in enumerate(COLS, 1):
            cell = ws.cell(1, ci, col)
            cell.font, cell.fill = HFONT, HFILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BDR
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        for ri, row in enumerate(df.itertuples(index=False), 2):
            g = str(getattr(row, "grade", "")).upper() if "grade" in df.columns else ""
            rfill = PatternFill("solid", fgColor=GRADE_CLR.get(g, "FFFFFF"))
            for ci, (col, _) in enumerate(COLS, 1):
                val = getattr(row, col, "") if col in df.columns else ""
                if str(val) in ("nan", "None", "NaN", "False"):
                    val = ""
                if str(val) == "True":
                    val = "✅"
                cell = ws.cell(ri, ci, str(val) if val != "" else "")
                cell.font, cell.fill, cell.border = DFONT, rfill, BDR
                cell.alignment = (
                    WRAP_A if col in WRAP_COLS else CTR_A if col in CTR_COLS else LFT_A
                )
            ws.row_dimensions[ri].height = 14 if col not in WRAP_COLS else 60

        # Tab hướng dẫn gửi email
        ws2 = wb.create_sheet("📋 Hướng dẫn gửi")
        guide = [
            ("HƯỚNG DẪN GỬI COLD EMAIL", ""),
            ("", ""),
            ("1. Chuẩn bị", ""),
            ("   - Dùng email cá nhân (tên thật), KHÔNG dùng info@", ""),
            ("   - Chữ ký đầy đủ: tên, số điện thoại, link lịch", ""),
            ("   - Warm up domain mới ít nhất 2 tuần trước khi gửi bulk", ""),
            ("", ""),
            ("2. Gửi theo batch", ""),
            ("   - Batch 1 (tuần 1): Hospitality + Logistics — 40 emails", ""),
            ("   - Batch 2 (tuần 2): Manufacturing + IT — 35 emails", ""),
            ("   - Batch 3 (tuần 3): Finance + Healthcare — 25 emails", ""),
            ("   - Gửi max 30 email/ngày để tránh spam filter", ""),
            ("", ""),
            ("3. Follow-up", ""),
            ("   - Ngày 4: Follow-up nếu chưa có reply (chỉ 2-3 câu)", ""),
            ("   - Ngày 7: Follow-up lần 2 — đổi angle (offer case study)", ""),
            ("   - Ngày 14: Final follow-up — ngắn gọn, lịch sự dừng", ""),
            ("", ""),
            ("4. Tracking", ""),
            ("   - Dùng Streak (Gmail) hoặc Mixmax để track open/click", ""),
            ("   - Note kết quả vào cột bên phải mỗi lead", ""),
            ("", ""),
            ("5. KPI mục tiêu", ""),
            ("   - Open rate: >40%", ""),
            ("   - Reply rate: >5-8%", ""),
            ("   - Meeting booked: 3-5 từ 100 email", ""),
        ]
        hf = Font(bold=True, name="Calibri", size=11)
        bf = Font(name="Calibri", size=10)
        for ri, (k, v) in enumerate(guide, 1):
            cell = ws2.cell(ri, 1, k)
            cell.font = (
                hf if k.startswith(("HƯỚNG", "1.", "2.", "3.", "4.", "5.")) else bf
            )
        ws2.column_dimensions["A"].width = 60

        wb.save(output_path)
        print(f"\n🎉 Xuất xong: {output_path}")
    except ImportError:
        df.to_csv(
            output_path.replace(".xlsx", ".csv"), index=False, encoding="utf-8-sig"
        )
        print(f"✅ Xuất CSV: {output_path.replace('.xlsx','.csv')}")


# ── MAIN ──────────────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  📧 Gen Cold Email Batch — Team Building")
    print("=" * 55)

    # 1. Load data
    if not os.path.exists(INPUT_FILE):
        # Thử tìm file top100
        candidates = [
            f for f in os.listdir(".") if "top100" in f and f.endswith(".xlsx")
        ]
        if candidates:
            input_path = candidates[0]
            print(f"⚠️  Dùng: {input_path}")
        else:
            print(f"❌ Không tìm thấy {INPUT_FILE}")
            print("   Chạy python filter_top100.py trước!")
            sys.exit(1)
    else:
        input_path = INPUT_FILE

    print(f"\n📂 Đọc: {input_path}")
    df = pd.read_excel(input_path, dtype=str).fillna("")
    leads = df.to_dict(orient="records")
    print(f"   → {len(leads)} leads")

    # 2. Render templates
    print(f"\n✉️  Render email templates...")
    render_all, render_email, SENDER = load_templates()
    leads_with_email = render_all(leads)

    # 3. AI personalize (optional)
    if USE_AI and AI_BATCH > 0:
        leads_with_email = ai_personalize_batch(leads_with_email, n=AI_BATCH)

    # 4. Stats
    with_body = sum(1 for l in leads_with_email if l.get("email_body"))
    by_industry = {}
    for l in leads_with_email:
        ind = l.get("industry", "?")
        by_industry[ind] = by_industry.get(ind, 0) + 1

    print(f"\n📊 Kết quả:")
    print(f"   Tổng leads: {len(leads_with_email)}")
    print(f"   Có email:   {with_body}")
    print(f"   Theo ngành: {by_industry}")

    # 5. Export
    df_out = pd.DataFrame(leads_with_email)
    export_with_emails(df_out, OUTPUT_FILE)

    print(f"\n📋 Bước tiếp:")
    print(f"   1. Mở {OUTPUT_FILE} — review từng email")
    print(f"   2. Chỉnh tên sender trong email_templates.py → SENDER dict")
    print(f"   3. Gửi batch đầu: Hospitality + Logistics (40 email/tuần 1)")
    print(f"   4. Track reply → follow-up sau 4 ngày")


if __name__ == "__main__":
    main()
