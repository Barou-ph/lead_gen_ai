"""
app.py — Lead Intelligence Tool · Streamlit Cloud Edition
Deploy: streamlit.io (free)

Chức năng:
  🔍 Filter  — upload Excel → lọc leads theo rule
  📧 Email   — xem cold email đã gen, copy nhanh
  ❌ Rejected — xem leads bị loại + lý do

Crawl vẫn chạy LOCAL bằng tool.py, export Excel, upload lên đây để filter.

Chạy local: streamlit run app.py
"""

import io, re
from datetime import datetime
import pandas as pd
import streamlit as st

# ── PAGE CONFIG ───────────────────────────────────────────────────────
st.set_page_config(
    page_title="Lead.AI",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────
st.markdown(
    """
<style>
/* Dark theme override */
[data-testid="stApp"] { background: #0b0d15; color: #e2e8f0; }
[data-testid="stSidebar"] { background: #12151f !important; border-right: 1px solid #1e2235; }
[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stSlider label,
[data-testid="stSidebar"] .stCheckbox label,
[data-testid="stSidebar"] .stTextInput label { color: #94a3b8 !important; font-size: .75rem !important; text-transform: uppercase; letter-spacing: .08em; }

/* Metric cards */
[data-testid="stMetric"] { background: #161929; border: 1px solid #1e2235; border-radius: 10px; padding: 14px 18px !important; }
[data-testid="stMetricLabel"] { color: #64748b !important; font-size: .72rem !important; text-transform: uppercase; letter-spacing: .08em; }
[data-testid="stMetricValue"] { color: #e2e8f0 !important; font-size: 1.6rem !important; font-weight: 800 !important; }

/* Buttons */
.stButton > button {
  background: linear-gradient(135deg,#2563eb,#1d4ed8) !important;
  color: #fff !important; border: none !important;
  border-radius: 7px !important; font-weight: 700 !important;
  transition: opacity .15s, transform .15s !important;
}
.stButton > button:hover { opacity: .88 !important; transform: translateY(-1px) !important; }

/* Download button */
.stDownloadButton > button {
  background: linear-gradient(135deg,#16a34a,#15803d) !important;
  color: #fff !important; border: none !important; border-radius: 7px !important; font-weight: 700 !important;
}

/* Dataframe */
[data-testid="stDataFrame"] { border: 1px solid #1e2235; border-radius: 8px; overflow: hidden; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] { background: #12151f; gap: 2px; border-bottom: 1px solid #1e2235; }
.stTabs [data-baseweb="tab"] { background: transparent; color: #64748b; border-radius: 6px 6px 0 0; font-size: .85rem; }
.stTabs [aria-selected="true"] { background: #161929 !important; color: #fff !important; border-bottom: 2px solid #3b82f6 !important; }

/* File uploader */
[data-testid="stFileUploader"] { background: #161929; border: 1.5px dashed #1e2235; border-radius: 8px; padding: 12px; }

/* Expander */
[data-testid="stExpander"] { background: #161929; border: 1px solid #1e2235; border-radius: 8px; }

/* Info/warning/success */
[data-testid="stAlert"] { border-radius: 8px !important; }

/* Copy button area */
.copy-box {
  background: #161929; border: 1px solid #1e2235; border-radius: 8px;
  padding: 14px; margin-bottom: 12px; font-family: 'Cascadia Code', 'Consolas', monospace;
  font-size: .76rem; color: #94a3b8; line-height: 1.7; white-space: pre-wrap;
  word-break: break-word;
}
.email-header { font-size: .82rem; font-weight: 700; color: #e2e8f0; margin-bottom: 4px; }
.email-meta { font-size: .7rem; color: #3b82f6; margin-bottom: 8px; }
.pain-tag { font-size: .68rem; color: #64748b; font-style: italic; margin-top: 6px; }
.grade-A { color: #eab308; font-weight: 800; }
.grade-B { color: #22c55e; font-weight: 800; }
.grade-C { color: #3b82f6; font-weight: 800; }
.grade-D { color: #ef4444; font-weight: 800; }
.pill {
  display: inline-block; padding: 2px 10px; border-radius: 20px;
  font-size: .68rem; font-weight: 700; margin-right: 4px;
}
.pill-red { background: #1c0f0f; border: 1px solid #3f1515; color: #ef4444; }
.pill-green { background: #0f1f12; border: 1px solid #14532d; color: #22c55e; }
.pill-blue { background: #0f1a2e; border: 1px solid #1e3a6a; color: #3b82f6; }
.pill-yellow { background: #1c1a0f; border: 1px solid #713f12; color: #eab308; }
</style>
""",
    unsafe_allow_html=True,
)

# ══════════════════════════════════════════════════════════════════════
#  FILTER CONSTANTS — self-contained, không import module ngoài
# ══════════════════════════════════════════════════════════════════════
PERSONAL_DOMAINS = {
    "gmail.com",
    "yahoo.com",
    "hotmail.com",
    "outlook.com",
    "icloud.com",
    "mail.com",
    "protonmail.com",
    "yandex.com",
    "live.com",
    "msn.com",
    "aol.com",
}

NOISE_DOMAINS = {
    # Báo chí VN
    "vnexpress.net",
    "tuoitre.vn",
    "thanhnien.vn",
    "thanhnien.com.vn",
    "dantri.com.vn",
    "congan.com.vn",
    "cand.com.vn",
    "suckhoedoisong.vn",
    "baochinhphu.vn",
    "vtv.vn",
    "baomoi.com",
    "zing.vn",
    "kenh14.vn",
    "soha.vn",
    "vietnamplus.vn",
    "nhandan.vn",
    "thuysanvietnam.com.vn",
    "visabatimes.com.vn",
    "vir.com.vn",
    "vietcetera.com",
    "genk.vn",
    "tienphong.vn",
    "laodong.vn",
    "nld.com.vn",
    # Báo quốc tế
    "techcrunch.com",
    "forbes.com",
    "bloomberg.com",
    "reuters.com",
    "bbc.com",
    "bbc.co.uk",
    "pcmag.com",
    "techradar.com",
    "zdnet.com",
    "dealstreetasia.com",
    "coinotag.com",
    # Job boards
    "topcv.vn",
    "itviec.com",
    "vietnamworks.com",
    "vieclam24h.vn",
    "jobstreet.com",
    "hoteljob.vn",
    "timviec365.vn",
    "glints.com",
    "jobthai.com",
    "wellfound.com",
    "freelancer.com",
    "upwork.com",
    "remotepeople.com",
    # Market research / analyst
    "statista.com",
    "crunchbase.com",
    "clutch.co",
    "goodfirms.co",
    "tracxn.com",
    "pitchbook.com",
    # Gov
    "gov.vn",
    "chinhphu.vn",
    "worldbank.org",
    "adb.org",
    # SaaS nước ngoài
    "smallpdf.com",
    "moovit.com",
    "shopify.com",
    "salesforce.com",
    "openai.com",
    "microsoft.com",
    "google.com",
    "amazon.com",
    "apple.com",
    "ibm.com",
    "kpmg.com",
    "ey.com",
    "deloitte.com",
    "pwc.com",
    "aon.com",
    "sider.ai",
    "equatorial.com",
    "vogue.com",
    "logitech.com",
    "agencyvn.com",
    "vietnix.vn",
    # E-commerce VN lớn
    "tiki.vn",
    "shopee.vn",
    "lazada.vn",
    "sendo.vn",
    "thegioididong.com",
    "fptshop.com.vn",
    "dienmayxanh.com",
    "mobileworld.com.vn",
    "bachlongmobile.com",
    # Tutorial
    "geeksforgeeks.org",
    "w3schools.com",
    "stackoverflow.com",
    "reddit.com",
    "quora.com",
    "wikipedia.org",
    "britannica.com",
    "investopedia.com",
    "wikihow.com",
    "vietjack.com",
    "thivien.net",
    "kahoot.com",
    # Games
    "poki.com",
    "crazygames.com",
    "bgames.com",
    "hahagames.com",
    "kizi.com",
    # Travel
    "tripadvisor.com",
    "booking.com",
    "agoda.com",
    "lonelyplanet.com",
    # Misc
    "medium.com",
    "substack.com",
    "blogspot.com",
    "wordpress.com",
    "linkedin.com",
    "trangvangvietnam.com",
    "yellowpages.vn",
    "yellowpages.com.vn",
    "softvn.vn",
    "blacksnetwork.net",
    "vietpedia.vn",
}

REJECT_DESC_PHRASES = [
    "latest news",
    "breaking news",
    "tin tức mới",
    "read more articles",
    "tin tức",
    "báo điện tử",
    "toà soạn",
    "tổng biên tập",
    "chuyên trang tin tức",
    "giấy phép mạng xã hội",
    "tin nóng",
    "cập nhật liên tục",
    "market size",
    "market share",
    "cagr",
    "buy report",
    "download report",
    "find companies",
    "list of companies",
    "top companies in",
    "danh sách công ty",
    "online games",
    "play free",
    "scholarship",
    "tuition",
    "admissions",
    "mục lục",
    "table of contents",
    "definition of",
    "learn what",
    "chuyên tổ chức du lịch",
    "công ty du lịch",
    "lữ hành và du lịch",
    "ã ",
    "á»",
    "ä ",
    "ã©",
]

BAD_EMAIL_PREFIXES = {
    "toasoan",
    "ads",
    "noreply",
    "no-reply",
    "unsubscribe",
    "bounce",
    "postmaster",
    "mailer-daemon",
    "donotreply",
    "newsletter",
    "webmaster",
    "abuse",
    "spam",
    "giaitrixahoi",
    "phonghanhchinh",
    "chimaivir",
    "bvmtw",
    "nckh",
}

BAD_EMAIL_PATTERNS = [
    r"sentry",
    r"wixpress",
    r"bug-report",
    r"^name@",
    r"^your@",
    r"^youname@",
    r"@email\.com$",
    r"@yourcompany\.com$",
    r"@example\.com$",
    r"^test@",
    r"^demo@",
    r"@company\.com$",
    r"@mail\.com$",
    r"^[a-f0-9]{20,}@",
    r"@robots\.net$",
    r"@trangvangvietnam\.com$",
    r"@yellowpages\.vn$",
    r"^email@congty\.vn$",
    r"^email@domain\.com$",
    r"@vnnic\.vn$",
    r"@sohu\.com$",
    r"@qq\.com$",
    r"@163\.com$",
    r"@126\.com$",
    r"@sina\.com$",
    r"^\d{5,}@",
    r"^user@",
    r"^example@",
]

BAD_EMAIL_DOMAINS_EXTRA = {
    "sentry.io",
    "wixpress.com",
    "example.com",
    "yourcompany.com",
    "robots.net",
    "schema.org",
    "w3.org",
    "cloudflare.com",
    "sohu.com",
    "qq.com",
    "163.com",
    "126.com",
    "sina.com",
    "trangvangvietnam.com",
    "yellowpages.vn",
    "vnnic.vn",
    "congty.vn",
    "domain.com",
    "brandinfo.biz",
    "yourweb.com.vn",
    "moovit.com",
    "exotheme.com",
}

FOREIGN_EMAIL_TLDS = {".fi", ".bd", ".lk", ".pk", ".th", ".cn", ".jp", ".kr", ".tw"}
VN_PHONE_RE = re.compile(r"(?:(?:\+84|0084|84)\s*)?(?:0[3-9]\d{8}|\b[3-9]\d{8}\b)")


# ══════════════════════════════════════════════════════════════════════
#  FILTER HELPERS
# ══════════════════════════════════════════════════════════════════════
def _get_domain(url):
    try:
        from urllib.parse import urlparse

        return urlparse(str(url)).netloc.lower().replace("www.", "")
    except:
        return ""


def _root_domain(url):
    try:
        parts = str(url).split("//")[-1].split("/")[0].split(".")
        return ".".join(parts[-2:]) if len(parts) >= 2 else str(url)
    except:
        return str(url)


def _is_noise_domain(url):
    if not url:
        return True
    dom = _get_domain(url)
    if dom in NOISE_DOMAINS:
        return True
    for nd in NOISE_DOMAINS:
        if dom.endswith("." + nd):
            return True
    return False


def _desc_ok(desc):
    if not desc or str(desc).strip() in ("", "nan", "None", "NaN"):
        return True
    low = str(desc).lower()
    return not any(p in low for p in REJECT_DESC_PHRASES)


def _has_vn_phone(phones):
    if not phones or str(phones).strip() in ("", "nan", "None", "NaN"):
        return False
    return bool(VN_PHONE_RE.search(str(phones)))


def _is_vn_lead(website, phones):
    dom = _get_domain(website)
    return dom.endswith(".vn") or ".vn." in dom or _has_vn_phone(str(phones))


def _is_bad_email_prefix(email):
    if not email or "@" not in email:
        return True
    return email.split("@")[0].lower().strip() in BAD_EMAIL_PREFIXES


def _is_bad_email(email):
    if not email or "@" not in email or len(email) > 100:
        return True
    low = email.lower().strip()
    domain = low.split("@")[-1]
    prefix = low.split("@")[0]
    if domain in BAD_EMAIL_DOMAINS_EXTRA:
        return True
    for tld in FOREIGN_EMAIL_TLDS:
        if domain.endswith(tld):
            return True
    if any(re.search(p, low) for p in BAD_EMAIL_PATTERNS):
        return True
    if len(prefix) > 50:
        return True
    if _is_bad_email_prefix(email):
        return True
    return False


def _email_type(email):
    if not email or "@" not in email:
        return "none"
    dom = email.split("@")[-1].lower().strip()
    return "personal" if dom in PERSONAL_DOMAINS else "company"


def _fix_encoding(text):
    if not text or str(text) in ("nan", "None", "NaN", ""):
        return ""
    try:
        return str(text).encode("latin-1").decode("utf-8")
    except:
        return str(text)


def _reject_reason(d, cfg):
    if _is_noise_domain(str(d.get("website", ""))):
        return "noise_domain"
    if not _desc_ok(str(d.get("description", ""))):
        return "bad_description"
    grade = str(d.get("grade", "")).strip().upper()
    if grade not in cfg["grades"]:
        return f"grade_{grade or 'missing'}"
    try:
        score = float(d.get("score", 0) or 0)
    except:
        score = 0
    if score < cfg["min_score"]:
        return f"score_{int(score)}"
    website = str(d.get("website", "")).strip()
    if not website or not website.startswith("http"):
        return "no_website"
    phones = str(d.get("phones", ""))
    if not _is_vn_lead(website, phones):
        return "non_vn_lead"
    if cfg["require_email"]:
        em = str(d.get("best_email", "") or d.get("emails", "") or "").strip()
        first = em.split(",")[0].strip() if em else ""
        if not first or _is_bad_email(first):
            return "bad_email"
        if _is_bad_email_prefix(first):
            return "bad_email_prefix"
    if cfg["require_phone"]:
        if not _has_vn_phone(phones):
            return "no_phone"
    if cfg["industries"]:
        ind = str(d.get("industry", "")).lower()
        if ind not in [i.lower() for i in cfg["industries"]]:
            return f"industry_{ind or '?'}"
    return ""


def apply_filter(df, cfg):
    passed, rejected = [], []
    seen_emails, seen_domains = set(), set()
    for _, row in df.iterrows():
        d = row.to_dict()
        reason = _reject_reason(d, cfg)
        if reason:
            rejected.append({**d, "reject_reason": reason})
            continue
        em = (
            str(d.get("best_email", "") or d.get("emails", "") or "")
            .split(",")[0]
            .strip()
            .lower()
        )
        dom = _root_domain(str(d.get("website", "")))
        if em and em not in ("", "nan", "none"):
            if em in seen_emails:
                rejected.append({**d, "reject_reason": "duplicate_email"})
                continue
            seen_emails.add(em)
        if dom:
            if dom in seen_domains:
                rejected.append({**d, "reject_reason": "duplicate_domain"})
                continue
            seen_domains.add(dom)
        passed.append(d)

    df_pass = pd.DataFrame(passed) if passed else pd.DataFrame()
    df_rej = pd.DataFrame(rejected) if rejected else pd.DataFrame()

    if not df_pass.empty:
        # Fix encoding
        if "description" in df_pass.columns:
            df_pass["description"] = df_pass["description"].apply(_fix_encoding)
            df_pass["description"] = df_pass["description"].apply(
                lambda d: "" if (str(d).count("á»") > 1 or str(d).count("Ã") > 2) else d
            )

        # Email type
        def _first_em(r):
            em = str(r.get("best_email", "") or r.get("emails", "") or "")
            return em.split(",")[0].strip() if em else ""

        df_pass["email_type"] = df_pass.apply(
            lambda r: _email_type(_first_em(r.to_dict())), axis=1
        )
        df_pass["has_phone"] = df_pass.get(
            "phones", pd.Series([""] * len(df_pass))
        ).apply(lambda p: bool(_has_vn_phone(str(p))))
        # Sort
        go = {"A": 0, "B": 1, "C": 2, "D": 3}
        df_pass["_g"] = df_pass.get("grade", pd.Series(["D"] * len(df_pass))).apply(
            lambda g: go.get(str(g).upper(), 3)
        )
        df_pass["_s"] = pd.to_numeric(df_pass.get("score", 0), errors="coerce").fillna(
            0
        )
        df_pass["_et"] = df_pass["email_type"].map(
            {"company": 0, "personal": 1, "none": 2, "unknown": 1}
        )
        df_pass.sort_values(
            ["_g", "_et", "_s"], ascending=[True, True, False], inplace=True
        )
        df_pass.drop(columns=["_g", "_s", "_et"], inplace=True, errors="ignore")
        df_pass.reset_index(drop=True, inplace=True)

    return df_pass, df_rej


# ══════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════
GRADE_COLORS = {"A": "C6EFCE", "B": "FFEB9C", "C": "DDEEFF", "D": "FFCCCC"}


def df_to_excel_bytes(df_pass, df_rej=None):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HFILL = PatternFill("solid", fgColor="1F4E79")
    HFONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    DFONT = Font(name="Calibri", size=9)
    THIN = Side(style="thin", color="CCCCCC")
    BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR_A = Alignment(horizontal="center", vertical="center")
    LFT_A = Alignment(horizontal="left", vertical="top", wrap_text=True)
    LFN_A = Alignment(horizontal="left", vertical="center")
    WRAP_COLS = {"description", "email_body", "email_subject", "pain_point", "emails"}
    CTR_COLS = {
        "grade",
        "score",
        "email_type",
        "industry",
        "has_phone",
        "should_contact",
    }

    def write_sheet(ws, df, title="Sheet"):
        if df is None or df.empty:
            ws["A1"] = "Không có dữ liệu."
            return
        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(1, ci, col)
            cell.font, cell.fill, cell.alignment, cell.border = HFONT, HFILL, CTR_A, BDR
            ws.column_dimensions[get_column_letter(ci)].width = min(
                max(len(col) * 2, 10), 60
            )
        ws.row_dimensions[1].height = 22
        for ri, row in enumerate(df.itertuples(index=False), 2):
            grade = str(
                getattr(row, "grade", "") if "grade" in df.columns else ""
            ).upper()
            rfill = PatternFill("solid", fgColor=GRADE_COLORS.get(grade, "FFFFFF"))
            for ci, col in enumerate(cols, 1):
                val = getattr(row, col, "") if col in df.columns else ""
                try:
                    if pd.isna(val):
                        val = ""
                except:
                    pass
                cell = ws.cell(ri, ci, str(val) if val != "" else "")
                cell.font, cell.fill, cell.border = DFONT, rfill, BDR
                cell.alignment = (
                    LFT_A if col in WRAP_COLS else (CTR_A if col in CTR_COLS else LFN_A)
                )
            ws.row_dimensions[ri].height = 14
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "✅ Leads"
    write_sheet(ws1, df_pass)
    if df_rej is not None and not df_rej.empty:
        ws2 = wb.create_sheet("❌ Rejected")
        write_sheet(ws2, df_rej)
    # Stats sheet
    ws3 = wb.create_sheet("📊 Stats")
    hf = Font(bold=True, name="Calibri", size=10)
    bf = Font(name="Calibri", size=9)
    rows = [
        ("Total passed", len(df_pass)),
        ("Total rejected", len(df_rej) if df_rej is not None else 0),
    ]
    if not df_pass.empty and "grade" in df_pass.columns:
        for g, c in df_pass["grade"].value_counts().items():
            rows.append((f"Grade {g}", c))
    if not df_pass.empty and "email_type" in df_pass.columns:
        for et, c in df_pass["email_type"].value_counts().items():
            rows.append((f"Email {et}", c))
    if df_rej is not None and not df_rej.empty and "reject_reason" in df_rej.columns:
        rows.append(("", ""))
        for r, c in df_rej["reject_reason"].value_counts().head(10).items():
            rows.append((f"Reject: {r}", c))
    for ri, (k, v) in enumerate(rows, 1):
        ws3.cell(ri, 1, str(k)).font = hf
        ws3.cell(ri, 2, str(v)).font = bf
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════
#  SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════
defaults = {
    "raw_df": None,
    "filter_passed": None,
    "filter_rejected": None,
    "last_upload": "",
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ══════════════════════════════════════════════════════════════════════
#  HEADER
# ══════════════════════════════════════════════════════════════════════
st.markdown(
    """
<div style="background:#12151f;border-bottom:1px solid #1e2235;padding:14px 24px;margin:-1rem -1rem 1.5rem -1rem;display:flex;align-items:center;gap:12px">
  <span style="font-size:1.3rem;font-weight:900;letter-spacing:-.5px;color:#fff">Lead<span style="color:#3b82f6">.</span>AI</span>
  <span style="font-size:.68rem;background:#1e2235;color:#64748b;padding:2px 9px;border-radius:20px">v5.5 · Cloud</span>
  <span style="font-size:.72rem;color:#64748b;margin-left:8px">🔍 Filter · 📧 Email · ❌ Rejected</span>
</div>
""",
    unsafe_allow_html=True,
)


# ══════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 📂 Upload Leads")
    st.caption("Export từ tool.py local → upload lên đây")

    uploaded = st.file_uploader(
        "Excel hoặc CSV",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
    )
    if uploaded and uploaded.name != st.session_state["last_upload"]:
        try:
            df = (
                pd.read_csv(uploaded, dtype=str).fillna("")
                if uploaded.name.endswith(".csv")
                else pd.read_excel(uploaded, dtype=str).fillna("")
            )
            st.session_state["raw_df"] = df
            st.session_state["filter_passed"] = None
            st.session_state["filter_rejected"] = None
            st.session_state["last_upload"] = uploaded.name
            st.success(f"✅ {uploaded.name} — **{len(df)}** rows")
        except Exception as e:
            st.error(f"Lỗi đọc file: {e}")

    if st.session_state["raw_df"] is not None:
        df_src = st.session_state["raw_df"]
        st.caption(f"{len(df_src)} rows đang load · upload file mới để thay")

    st.markdown("---")
    st.markdown("### ⚙️ Filter Config")

    min_score = st.slider("Min Score", 0, 50, 20, help="Score tối thiểu để pass")

    grades_sel = st.multiselect(
        "Grades",
        ["A", "B", "C", "D"],
        default=["A", "B"],
        help="Chỉ lấy Grade A+B cho batch đầu",
    )

    require_email = st.checkbox("Cần Company Email", value=True)
    require_phone = st.checkbox("Cần Phone VN", value=False)

    ind_input = st.text_input(
        "Industry filter",
        placeholder="hospitality, logistics... (bỏ trống = tất cả)",
        help="Ngăn cách bằng dấu phẩy",
    )

    run_btn = st.button(
        "🔍 Lọc Leads",
        type="primary",
        use_container_width=True,
        disabled=st.session_state["raw_df"] is None,
    )

    if run_btn:
        df_src = st.session_state["raw_df"]
        cfg = {
            "min_score": min_score,
            "grades": [g.upper() for g in grades_sel],
            "require_email": require_email,
            "require_phone": require_phone,
            "industries": [x.strip() for x in ind_input.split(",") if x.strip()],
        }
        with st.spinner("Đang lọc..."):
            df_p, df_r = apply_filter(df_src.copy(), cfg)
        st.session_state["filter_passed"] = df_p
        st.session_state["filter_rejected"] = df_r
        st.success(f"✅ {len(df_p)} leads | ❌ {len(df_r)} rejected")

    st.markdown("---")
    st.markdown("### 📋 Hướng dẫn")
    st.markdown("""
**Workflow:**
1. Chạy `tool.py` local → crawl leads
2. Export Excel từ tool
3. Upload Excel lên đây
4. Lọc → xem email → export

**Gửi email tuần 1:**
- Hospitality + Logistics
- Grade A+B, score ≥ 20
- Max 30 email/ngày
""")


# ══════════════════════════════════════════════════════════════════════
#  MAIN TABS
# ══════════════════════════════════════════════════════════════════════
df_pass = st.session_state.get("filter_passed")
df_rej = st.session_state.get("filter_rejected")

tab_filter, tab_email, tab_rejected = st.tabs(
    ["🔍 Filter Results", "📧 Cold Emails", "❌ Rejected"]
)

# ─── TAB: FILTER ─────────────────────────────────────────────────────
with tab_filter:
    if df_pass is not None:
        # Stats row
        gc = (
            df_pass["grade"].value_counts().to_dict()
            if "grade" in df_pass.columns
            else {}
        )
        etc = (
            df_pass["email_type"].value_counts().to_dict()
            if "email_type" in df_pass.columns
            else {}
        )
        sc = (
            df_pass["should_contact"].apply(lambda x: x == "✓").sum()
            if "should_contact" in df_pass.columns
            else 0
        )

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("✅ Total", len(df_pass))
        c2.metric("🏆 Grade A", gc.get("A", 0))
        c3.metric("🥈 Grade B", gc.get("B", 0))
        c4.metric("📧 Company Email", etc.get("company", 0))
        c5.metric(
            "🎯 Gửi ngay",
            (
                sc
                if sc
                else (
                    len(df_pass[df_pass.get("grade", "") == "A"])
                    if "grade" in df_pass.columns
                    else "—"
                )
            ),
        )

        st.markdown("")

        # Export button
        col_exp, col_space = st.columns([1, 3])
        with col_exp:
            excel_bytes = df_to_excel_bytes(df_pass, df_rej)
            st.download_button(
                "⬇️ Export Excel",
                data=excel_bytes,
                file_name=f"leads_filtered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # Display columns
        display_cols = [
            c
            for c in [
                "grade",
                "score",
                "website",
                "best_email",
                "email_type",
                "phones",
                "industry",
                "field",
                "should_contact",
                "description",
            ]
            if c in df_pass.columns
        ]

        # Style dataframe
        def _style_grade(val):
            colors = {"A": "#eab308", "B": "#22c55e", "C": "#3b82f6", "D": "#ef4444"}
            return f"color:{colors.get(str(val).upper(),'#e2e8f0')};font-weight:800"

        st.dataframe(
            df_pass[display_cols] if display_cols else df_pass,
            use_container_width=True,
            height=520,
            column_config={
                "website": st.column_config.LinkColumn("Website", width=200),
                "grade": st.column_config.TextColumn("Grade", width=60),
                "score": st.column_config.NumberColumn("Score", width=60),
                "best_email": st.column_config.TextColumn("Email", width=200),
                "email_type": st.column_config.TextColumn("Type", width=80),
                "phones": st.column_config.TextColumn("Phones", width=160),
                "industry": st.column_config.TextColumn("Industry", width=100),
                "description": st.column_config.TextColumn("Description", width=300),
                "should_contact": st.column_config.TextColumn("✓", width=40),
            },
        )

    elif st.session_state["raw_df"] is not None:
        st.info(
            f"Có **{len(st.session_state['raw_df'])}** rows — nhấn **🔍 Lọc Leads** ở sidebar để filter."
        )
        st.dataframe(st.session_state["raw_df"].head(10), use_container_width=True)
    else:
        st.markdown(
            """
<div style="text-align:center;padding:60px 20px;color:#64748b">
  <div style="font-size:3rem;margin-bottom:16px">📂</div>
  <div style="font-size:1rem;font-weight:600;color:#94a3b8;margin-bottom:8px">Upload file để bắt đầu</div>
  <div style="font-size:.82rem">Export từ tool.py local → upload Excel lên sidebar</div>
</div>
""",
            unsafe_allow_html=True,
        )


# ─── TAB: EMAIL ──────────────────────────────────────────────────────
with tab_email:
    if df_pass is not None and not df_pass.empty:
        has_email_col = "email_body" in df_pass.columns
        has_subject = "email_subject" in df_pass.columns

        if has_email_col:
            df_em = df_pass[
                df_pass["email_body"].apply(
                    lambda x: bool(str(x).strip() and str(x) not in ("", "nan", "None"))
                )
            ].copy()
        else:
            df_em = pd.DataFrame()

        if df_em.empty:
            st.info(
                "Leads chưa có email body. Gen email bằng `gen_email_batch.py` local rồi upload lại."
            )

            # Hiển thị preview template
            st.markdown("### 📝 Preview email template (hospitality)")
            st.code(
                """Subject: Pearl River Hotel — giải pháp gắn kết nhân viên mùa cao điểm

Chào Anh/Chị,

Mình là Minh Anh từ Vibe Team Building.

Ngành khách sạn / nhà hàng đang đối mặt với tỷ lệ nghỉ việc cao nhất...
[Chạy gen_email_batch.py để gen đầy đủ]""",
                language=None,
            )
        else:
            st.markdown(f"**{len(df_em)}/{len(df_pass)} leads đã có email**")

            # Filter by industry
            if "industry" in df_em.columns:
                industries = ["Tất cả"] + sorted(
                    df_em["industry"].dropna().unique().tolist()
                )
                sel_ind = st.selectbox(
                    "Lọc theo ngành", industries, label_visibility="collapsed"
                )
                if sel_ind != "Tất cả":
                    df_em = df_em[df_em["industry"] == sel_ind]

            st.markdown("")

            for _, row in df_em.iterrows():
                try:
                    from urllib.parse import urlparse

                    site = urlparse(str(row.get("website", ""))).netloc.replace(
                        "www.", ""
                    )
                except:
                    site = str(row.get("website", ""))

                grade = str(row.get("grade", "?"))
                email = str(row.get("best_email", ""))
                subject = str(row.get("email_subject", "(no subject)"))
                body = str(row.get("email_body", ""))
                pain = str(row.get("pain_point", ""))
                industry = str(row.get("industry", ""))

                grade_cls = f"grade-{grade}" if grade in "ABCD" else ""

                with st.expander(
                    f"**{site}** · `{email}` · Grade {grade}", expanded=False
                ):
                    col_info, col_copy = st.columns([3, 1])
                    with col_info:
                        st.markdown(
                            f"""
<div class="email-header">📧 {subject}</div>
<div class="email-meta">→ {email} &nbsp;·&nbsp; {industry} &nbsp;·&nbsp; <span class="{grade_cls}">Grade {grade}</span></div>
""",
                            unsafe_allow_html=True,
                        )

                    with col_copy:
                        # Copy subject
                        st.text_area(
                            "Subject",
                            subject,
                            height=60,
                            key=f"subj_{email}",
                            label_visibility="collapsed",
                        )

                    st.text_area(
                        "Email body",
                        body,
                        height=300,
                        key=f"body_{email}",
                        label_visibility="collapsed",
                        help="Ctrl+A → Ctrl+C để copy toàn bộ",
                    )
                    if pain and pain not in ("nan", "None", ""):
                        st.caption(f"💡 Pain point: {pain}")
    else:
        st.info("Filter leads trước ở tab **🔍 Filter Results**.")


# ─── TAB: REJECTED ───────────────────────────────────────────────────
with tab_rejected:
    if df_rej is not None and not df_rej.empty:
        # Reject reason pills
        if "reject_reason" in df_rej.columns:
            rc = df_rej["reject_reason"].value_counts().head(10)
            st.markdown("**Lý do bị loại:**")
            pills_html = " ".join(
                f'<span class="pill pill-red"><b>{cnt}</b> {reason}</span>'
                for reason, cnt in rc.items()
            )
            st.markdown(
                f'<div style="margin-bottom:16px">{pills_html}</div>',
                unsafe_allow_html=True,
            )

        c1, c2 = st.columns([1, 3])
        with c1:
            # Export rejected
            rej_bytes = df_to_excel_bytes(df_rej)
            st.download_button(
                "⬇️ Export Rejected",
                data=rej_bytes,
                file_name=f"leads_rejected_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        show_cols = [
            c
            for c in [
                "grade",
                "score",
                "website",
                "reject_reason",
                "industry",
                "field",
                "phones",
            ]
            if c in df_rej.columns
        ]
        st.dataframe(
            df_rej[show_cols] if show_cols else df_rej,
            use_container_width=True,
            height=500,
            column_config={
                "website": st.column_config.LinkColumn("Website"),
                "reject_reason": st.column_config.TextColumn("Lý do", width=150),
                "grade": st.column_config.TextColumn("Grade", width=60),
                "score": st.column_config.NumberColumn("Score", width=60),
            },
        )
    else:
        st.markdown(
            """
<div style="text-align:center;padding:40px 20px;color:#64748b">
  <div style="font-size:2rem;margin-bottom:10px">✅</div>
  <div>Chưa có dữ liệu — chạy Filter trước.</div>
</div>
""",
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════
#  FOOTER
# ══════════════════════════════════════════════════════════════════════
st.markdown(
    """
<div style="text-align:center;color:#374151;font-size:.7rem;margin-top:2rem;padding-top:1rem;border-top:1px solid #1e2235">
  Lead.AI v5.5 · Filter &amp; Email Viewer · Crawl chạy local bằng <code>tool.py</code>
</div>
""",
    unsafe_allow_html=True,
)
