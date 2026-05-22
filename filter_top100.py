"""
filter_top100.py — Lọc top 100 leads từ leads_ALL_v1.xlsx
Chạy: python filter_top100.py
Output: top100_leads.xlsx
"""

import os, sys
import pandas as pd
from urllib.parse import urlparse

# Import từ filter.py (nguồn chân lý) để đồng bộ luật lọc
from filter import filter_url, is_valid_email, validate_description, is_vietnam_lead

# ── CONFIG ────────────────────────────────────────────────────────────
INPUT_FILE = "leads_ALL_v1.xlsx"  # hoặc leads_ALL.xlsx
OUTPUT_FILE = "top100_leads.xlsx"
TARGET = 100
PRIORITY_INDUSTRIES = [
    "hospitality",
    "logistics",
    "manufacturing",
    "it",
    "finance",
    "healthcare",
]

# ── HELPERS ───────────────────────────────────────────────────────────
def get_domain(url):
    try:
        return urlparse(str(url)).netloc.lower().replace("www.", "")
    except:
        return ""

def get_best_email(row):
    for col in ["best_email", "emails"]:
        if col in row and str(row[col]).strip() not in ("", "nan", "None", "NaN"):
            return str(row[col]).split(",")[0].strip()
    return ""

def ind_priority(industry):
    try:
        return PRIORITY_INDUSTRIES.index(str(industry).lower())
    except:
        return 99

def email_type(email):
    if not email or "@" not in str(email):
        return "none"
    dom = str(email).split("@")[-1].lower()
    from filter import BAD_EMAIL_DOMAINS
    # Đánh dấu sơ bộ là personal hay company
    if dom in {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "icloud.com"}:
        return "personal"
    return "company"

# ── MAIN ──────────────────────────────────────────────────────────────
def main():
    # Tìm file input
    candidates = [INPUT_FILE, "leads_ALL.xlsx", "leads_ALL_v1.xlsx"]
    input_path = None
    for c in candidates:
        if os.path.exists(c):
            input_path = c
            break
    if not input_path:
        all_xlsx = [
            f for f in os.listdir(".") if f.endswith(".xlsx") and "lead" in f.lower()
        ]
        if all_xlsx:
            input_path = all_xlsx[0]
            print(f"⚠️  Không tìm thấy {INPUT_FILE}, dùng: {input_path}")
        else:
            print(f"❌ Không tìm thấy file Excel. Hãy đặt file vào cùng thư mục.")
            sys.exit(1)

    print(f"📂 Đọc: {input_path}")
    df = pd.read_excel(input_path, dtype=str).fillna("")
    print(f"   → {len(df)} leads tổng")

    rejected = []
    passed = []
    seen_emails = set()
    seen_domains = set()

    for _, row in df.iterrows():
        d = row.to_dict()

        # 1. Website check
        ws = str(d.get("website", "")).strip()
        if not ws or not ws.startswith("http"):
            rejected.append({**d, "_reason": "no_website"})
            continue
            
        keep, reason = filter_url(ws)
        if not keep:
            rejected.append({**d, "_reason": reason})
            continue

        # 2. Tầng 5 - Vietnam Lead Check (Domain VN hoặc Phone VN)
        phones = str(d.get("phones", "")).strip()
        if not is_vietnam_lead(ws, phones):
            rejected.append({**d, "_reason": "non_vn_lead"})
            continue

        # 3. Description check
        desc = str(d.get("description", "")).strip()
        is_desc_ok, desc_reason = validate_description(desc)
        if not is_desc_ok:
            rejected.append({**d, "_reason": desc_reason})
            continue

        # 4. Email check (Cho phép khuyết email, nhưng nếu có phải là email chuẩn)
        em = get_best_email(d)
        if em and not is_valid_email(em):
            rejected.append({**d, "_reason": "bad_email"})
            continue

        # Dedup by domain
        dom = get_domain(ws)
        if dom and dom in seen_domains:
            rejected.append({**d, "_reason": "dup_domain"})
            continue
        if dom:
            seen_domains.add(dom)

        # Dedup by email (nếu có email)
        if em:
            em_low = em.lower()
            if em_low in seen_emails:
                rejected.append({**d, "_reason": "dup_email"})
                continue
            seen_emails.add(em_low)

        # Tag email type
        d["email_type"] = email_type(em)
        d["best_email_clean"] = em
        
        try:
            score = float(d.get("score", 0) or 0)
        except:
            score = 0
            
        grade = str(d.get("grade", "")).strip().upper()

        d["_score_num"] = score
        d["_ind_pri"] = ind_priority(d.get("industry", ""))
        d["_grade_num"] = {"A": 0, "B": 1}.get(grade, 2)
        d["_em_type_n"] = 0 if d["email_type"] == "company" else (1 if d["email_type"] == "personal" else 2)
        
        passed.append(d)

    print(f"\n✅ Sau filter cứng: {len(passed)} leads thật | ❌ Rejected: {len(rejected)}")

    # Sort: A > B | company email > personal > none | score desc | industry priority
    passed.sort(
        key=lambda x: (
            x["_grade_num"],
            x["_em_type_n"],
            x["_ind_pri"],
            -x["_score_num"],
        )
    )

    # Lấy top 100
    top = passed[:TARGET]

    # Cleanup sort keys
    for d in top:
        for k in ["_score_num", "_ind_pri", "_grade_num", "_em_type_n"]:
            d.pop(k, None)

    df_top = pd.DataFrame(top)
    df_rej = pd.DataFrame(rejected)

    # ── Stats ──────────────────────────────────────────────────────────
    print(f"\n📊 TOP {len(top)} LEADS:")
    if "grade" in df_top.columns and not df_top.empty:
        print(f"   Grade: {df_top['grade'].value_counts().to_dict()}")
    if "industry" in df_top.columns and not df_top.empty:
        print(f"   Industry:\n   {df_top['industry'].value_counts().to_dict()}")
    if "email_type" in df_top.columns and not df_top.empty:
        print(f"   Email type: {df_top['email_type'].value_counts().to_dict()}")

    reject_reasons = (
        df_rej["_reason"].value_counts() if not df_rej.empty and "_reason" in df_rej.columns else {}
    )
    print(f"\n❌ Reject reasons: {reject_reasons.to_dict() if not reject_reasons.empty else {}}")

    # ── Xuất Excel đẹp ────────────────────────────────────────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        GRADE_CLR = {"A": "FFD700", "B": "90EE90", "C": "ADD8E6"}
        THIN = Side(style="thin", color="CCCCCC")
        BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        HFILL = PatternFill("solid", fgColor="1F4E79")
        HFONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        DFONT = Font(name="Calibri", size=9)
        CTR_A = Alignment(horizontal="center", vertical="center")
        LFT_A = Alignment(horizontal="left", vertical="center", wrap_text=False)
        WRAP_A = Alignment(horizontal="left", vertical="top", wrap_text=True)

        COLS = [
            ("grade", 7),
            ("score", 7),
            ("best_email_clean", 34),
            ("email_type", 12),
            ("phones", 22),
            ("website", 42),
            ("field", 18),
            ("industry", 14),
            ("description", 55),
            ("session", 20),
        ]
        CTR_COLS = {"grade", "score", "email_type", "industry"}
        WRAP_COLS = {"description"}

        wb = Workbook()
        ws = wb.active
        ws.title = f"Top {len(top)} Leads"

        # Header
        for ci, (col, width) in enumerate(COLS, 1):
            cell = ws.cell(1, ci, col)
            cell.font, cell.fill = HFONT, HFILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BDR
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        # Rows
        for ri, row in enumerate(df_top.itertuples(index=False), 2):
            g = str(getattr(row, "grade", "")).upper()
            rfill = PatternFill("solid", fgColor=GRADE_CLR.get(g, "FFFFFF"))
            for ci, (col, _) in enumerate(COLS, 1):
                val = getattr(row, col, "") if col in df_top.columns else ""
                if str(val) in ("nan", "None", "NaN"):
                    val = ""
                cell = ws.cell(ri, ci, str(val) if val != "" else "")
                cell.font, cell.fill, cell.border = DFONT, rfill, BDR
                cell.alignment = (
                    WRAP_A if col in WRAP_COLS else CTR_A if col in CTR_COLS else LFT_A
                )
            ws.row_dimensions[ri].height = 15

        # Stats sheet
        ws2 = wb.create_sheet("Stats")
        hf = Font(bold=True, name="Calibri", size=10)
        bf = Font(name="Calibri", size=9)
        rows_stats = [
            ("Total input", len(df)),
            ("Passed filter", len(passed)),
            ("Top selected", len(top)),
            ("Rejected", len(rejected)),
            ("", ""),
            ("=== Grade ===", ""),
        ]
        if "grade" in df_top.columns and not df_top.empty:
            for g, c in df_top["grade"].value_counts().items():
                rows_stats.append((f"Grade {g}", c))
        rows_stats += [("", ""), ("=== Industry ===", "")]
        if "industry" in df_top.columns and not df_top.empty:
            for ind, c in df_top["industry"].value_counts().items():
                rows_stats.append((ind, c))
        rows_stats += [("", ""), ("=== Email Type ===", "")]
        if "email_type" in df_top.columns and not df_top.empty:
            for et, c in df_top["email_type"].value_counts().items():
                rows_stats.append((et, c))
        rows_stats += [("", ""), ("=== Reject Reasons ===", "")]
        if not reject_reasons.empty:
            for r, c in reject_reasons.items():
                rows_stats.append((r, c))
        for ri, (k, v) in enumerate(rows_stats, 1):
            ws2.cell(ri, 1, str(k)).font = hf if str(k).startswith("===") else bf
            ws2.cell(ri, 2, str(v) if v != "" else "").font = bf
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 12

        wb.save(OUTPUT_FILE)
        print(f"\n🎉 Xuất xong: {OUTPUT_FILE} ({len(top)} leads)")

    except ImportError:
        # Fallback: xuất CSV
        df_top.to_csv(
            OUTPUT_FILE.replace(".xlsx", ".csv"), index=False, encoding="utf-8-sig"
        )
        print(
            f"\n✅ Xuất CSV (cài openpyxl để ra .xlsx): {OUTPUT_FILE.replace('.xlsx','.csv')}"
        )

if __name__ == "__main__":
    main()
