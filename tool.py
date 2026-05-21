"""
tool.py — Lead Intelligence Tool v5 (All-in-One)
Tích hợp crawler.py (YellowPages) + filter + cold email vào 1 app.

Chạy: python tool.py
Mở:   http://localhost:5000

Tab 🕷️  Crawl    → chọn ngành → crawl YellowPages → leads
Tab 🔍  Filter   → upload Excel hoặc dùng kết quả crawl → lọc
Tab 📧  Email    → gen cold email bằng AI → copy gửi
Tab ❌  Rejected → xem leads bị loại + lý do
"""

import os, json, threading, time, re, random, base64
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, send_file
import pandas as pd

app = Flask(__name__)

# ══════════════════════════════════════════════════════════════════════
#  SHARED STATE
# ══════════════════════════════════════════════════════════════════════
_lock = threading.Lock()

_crawl = {
    "status": "idle",  # idle | running | stopping | done | error
    "logs": [],
    "_log_cursor": 0,
    "leads": [],
    "session": "",
    "total_urls": 0,
    "done_urls": 0,
    "phase": "",
}
_gen = {
    "status": "idle",
    "logs": [],
    "_log_cursor": 0,
    "progress": 0,
    "total": 0,
    "leads": [],
}
_filter_state = {
    "df_raw": None,
    "df_filtered": None,
    "df_rejected": None,
}

# ══════════════════════════════════════════════════════════════════════
#  INDUSTRY META — hiển thị trên UI
# ══════════════════════════════════════════════════════════════════════
INDUSTRY_META = {
    "manufacturing": {
        "label": "🏭 Manufacturing",
        "desc": "Nhà máy, sản xuất, cơ khí",
        "color": "#e67e22",
        "why": "Ít noise nhất — yellowpages có danh sách nhà máy thật 100%",
    },
    "logistics": {
        "label": "🚚 Logistics",
        "desc": "Vận tải, kho bãi, freight",
        "color": "#2980b9",
        "why": "Công ty logistics VN nhiều, dễ tìm email sales",
    },
    "hospitality": {
        "label": "🏨 Hospitality",
        "desc": "Khách sạn, nhà hàng, du lịch",
        "color": "#8e44ad",
        "why": "Team building là nhu cầu cốt lõi, budget tốt",
    },
    "healthcare": {
        "label": "🏥 Healthcare",
        "desc": "Phòng khám, dược phẩm, spa",
        "color": "#27ae60",
        "why": "Private clinic + pharma — ngân sách lớn, hay tổ chức team building",
    },
    "realestate": {
        "label": "🏗️ Real Estate",
        "desc": "BĐS, xây dựng, nội thất",
        "color": "#c0392b",
        "why": "Sales team lớn, incentive trip là nhu cầu thường xuyên",
    },
    "finance": {
        "label": "💰 Finance",
        "desc": "Tài chính, bảo hiểm, kế toán",
        "color": "#f39c12",
        "why": "Budget lớn, chuộng team building & retreat cao cấp",
    },
    "retail": {
        "label": "🛒 Retail",
        "desc": "Bán lẻ, chuỗi, siêu thị",
        "color": "#16a085",
        "why": "Chuỗi nhiều nhân viên, cần gắn kết thường xuyên",
    },
    "education": {
        "label": "📚 Education",
        "desc": "Trường, trung tâm đào tạo",
        "color": "#2c3e50",
        "why": "Trung tâm tư thục có ngân sách, cần team building cho staff",
    },
    "it": {
        "label": "💻 IT / Tech",
        "desc": "Phần mềm, outsourcing, agency",
        "color": "#1abc9c",
        "why": "Công ty IT VN nhiều, developer burnout = nhu cầu team building cao",
    },
}

# ══════════════════════════════════════════════════════════════════════
#  YELLOWPAGES CATEGORIES — đồng bộ từ crawler.py
# ══════════════════════════════════════════════════════════════════════
YP_CATEGORIES = {
    "manufacturing": [
        "https://www.yellowpages.vn/cls/152060/co-khi----gia-cong-va-che-tao.html",
        "https://www.yellowpages.vn/cls/111010/do-go-noi-that---san-xuat-va-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/47910/hoa-chat---san-xuat,-nhap-khau-va-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/186060/nhua---cac-cong-ty-nhua.html",
        "https://www.yellowpages.vn/cls/174170/bao-bi---nha-san-xuat-va-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/225100/thep---cong-ty-thep-(san-xuat,-kinh-doanh).html",
        "https://www.yellowpages.vn/cls/64610/thuc-pham---san-xuat-va-che-bien.html",
        "https://www.yellowpages.vn/cls/97160/vai-soi---san-xuat-va-kinh-doanh.html",
    ],
    "hospitality": [
        "https://www.yellowpages.vn/cls/127160/khach-san.html",
        "https://www.yellowpages.vn/cls/200710/nha-hang.html",
        "https://www.yellowpages.vn/cls/51810/du-lich---cong-ty-lu-hanh-va-du-lich.html",
        "https://www.yellowpages.vn/cls/106010/thuc-pham---cung-cap-thuc-pham,-cong-ty-thuc-pham.html",
        "https://www.yellowpages.vn/cls/56010/cafe---quan-ca-phe-va-tra-sua.html",
    ],
    "logistics": [
        "https://www.yellowpages.vn/cls/246160/van-tai---cong-ty-van-tai-va-dai-ly-van-tai.html",
        "https://www.yellowpages.vn/cls/485215/van-tai-duong-bo.html",
        "https://www.yellowpages.vn/cls/213810/van-tai-bien.html",
        "https://www.yellowpages.vn/cls/68660/van-tai-container.html",
        "https://www.yellowpages.vn/cls/130610/giao-nhan-hang-hoa---chuyen-phat.html",
        "https://www.yellowpages.vn/cls/48310/kho-bai---cho-thue-va-quan-ly-kho.html",
        "https://www.yellowpages.vn/cls/135510/logistics---cong-ty-logistics.html",
    ],
    "healthcare": [
        "https://www.yellowpages.vn/cls/180660/duoc-pham---cong-ty-duoc-pham.html",
        "https://www.yellowpages.vn/cls/152660/y-te---benh-vien-va-co-so-chuyen-khoa.html",
        "https://www.yellowpages.vn/cls/424960/phong-kham-da-khoa.html",
        "https://www.yellowpages.vn/cls/77080/phong-kham-nha-khoa.html",
        "https://www.yellowpages.vn/cls/152560/thiet-bi-y-te---san-xuat,-kinh-doanh.html",
        "https://www.yellowpages.vn/cls/126660/spa---tham-my-vien.html",
    ],
    "realestate": [
        "https://www.yellowpages.vn/cls/192550/bat-dong-san---cac-cong-ty-bat-dong-san.html",
        "https://www.yellowpages.vn/cls/37210/xay-dung---cong-ty-xay-dung.html",
        "https://www.yellowpages.vn/cls/420940/do-noi-that---thiet-ke-va-san-xuat.html",
        "https://www.yellowpages.vn/cls/197660/bat-dong-san---quan-ly-va-tu-van-bat-dong-san.html",
        "https://www.yellowpages.vn/cls/169260/van-phong---cho-thue-van-phong-(tron-goi,-ao,-co---working-space).html",
    ],
    "finance": [
        "https://www.yellowpages.vn/cls/131560/bao-hiem---cong-ty-bao-hiem.html",
        "https://www.yellowpages.vn/cls/10310/ke-toan-va-kiem-toan.html",
        "https://www.yellowpages.vn/cls/488245/ke-toan-thue---dich-vu-ke-toan-thue.html",
        "https://www.yellowpages.vn/cls/100310/tai-chinh---cac-cong-ty-tai-chinh.html",
        "https://www.yellowpages.vn/cls/38760/tu-van-quan-ly-doanh-nghiep.html",
    ],
    "retail": [
        "https://www.yellowpages.vn/cls/229210/sieu-thi.html",
        "https://www.yellowpages.vn/cls/225060/may-mac---quan-ao-thoi-trang.html",
        "https://www.yellowpages.vn/cls/116010/thuc-pham---san-xuat-va-phan-phoi.html",
        "https://www.yellowpages.vn/cls/147410/sieu-thi---trung-tam-thuong-mai.html",
    ],
    "education": [
        "https://www.yellowpages.vn/cls/245665/dao-tao---cac-cong-ty-dao-tao.html",
        "https://www.yellowpages.vn/cls/335515/dao-tao---ngoai-ngu.html",
        "https://www.yellowpages.vn/cls/487018/trung-tam-dao-tao.html",
        "https://www.yellowpages.vn/cls/140360/truong-mam-non---truong-mau-giao.html",
        "https://www.yellowpages.vn/cls/164810/ngoai-ngu---trung-tam-ngoai-ngu.html",
    ],
    "it": [
        "https://www.yellowpages.vn/cls/66030/phan-mem.html",
        "https://www.yellowpages.vn/cls/226260/cong-nghe-thong-tin---it-services.html",
        "https://www.yellowpages.vn/cls/166210/thiet-ke-web---cong-ty-thiet-ke-website.html",
        "https://www.yellowpages.vn/cls/231510/digital-marketing---agency.html",
        "https://www.yellowpages.vn/cls/373705/phan-mem-quan-tri.html",
        "https://www.yellowpages.vn/cls/493201/phan-mem---dich-vu-gia-cong-phan-mem.html",
    ],
}

SKIP_DOMAINS_YP = {
    "yellowpages.vn",
    "yellowpages.com.vn",
    "google.com",
    "facebook.com",
    "zalo.me",
    "youtube.com",
    "twitter.com",
    "instagram.com",
    "tiktok.com",
    "linkedin.com",
    "apple.com",
    "microsoft.com",
    "maps.google.com",
    "trangvangvietnam.com",
    "signup.trangvangvietnam.com",
    "tiki.vn",
    "shopee.vn",
    "lazada.vn",
    "sendo.vn",
    "thegioididong.com",
    "fptshop.com.vn",
    "dienmayxanh.com",
    "vnexpress.net",
    "tuoitre.vn",
    "thanhnien.vn",
    "dantri.com.vn",
    "wikipedia.org",
    "topcv.vn",
    "itviec.com",
    "vietnamworks.com",
    "vieclam24h.vn",
}

# ══════════════════════════════════════════════════════════════════════
#  FILTER BLACKLISTS (sync với filter.py)
# ══════════════════════════════════════════════════════════════════════
PERSONAL_DOMAINS = {
    "gmail.com",
    "yahoo.com",
    "hotmail.com",
    "outlook.com",
    "icloud.com",
    "mail.com",
    "protonmail.com",
    "yandex.com",
    "live.com",
    "msn.com",
    "aol.com",
}

NOISE_DOMAINS = {
    "vnexpress.net",
    "tuoitre.vn",
    "thanhnien.vn",
    "dantri.com.vn",
    "vietcetera.com",
    "kr-asia.com",
    "e27.co",
    "techcrunch.com",
    "forbes.com",
    "bloomberg.com",
    "reuters.com",
    "dealstreetasia.com",
    "asiapevc.com",
    "privateequityinternational.com",
    "businessnewsasia.com",
    "techwireasia.com",
    "livebitcoinnews.com",
    "hivelife.com",
    "iglu.net",
    "thefintechmag.com",
    "vuihoc.vn",
    "vietnamfinance.vn",
    "robots.net",
    "startuprise.org",
    "coinotag.com",
    "bestdevops.com",
    "eventsnewsasia.com",
    "cri-report.com",
    "edisongroup.com",
    "pcmag.com",
    "techradar.com",
    "zdnet.com",
    "bbc.com",
    "bbc.co.uk",
    "statista.com",
    "mordorintelligence.com",
    "expertmarketresearch.com",
    "itif.org",
    "growyourbusiness.org",
    "clutch.co",
    "goodfirms.co",
    "tracxn.com",
    "crunchbase.com",
    "zoominfo.com",
    "topon.tech",
    "investasian.com",
    "topmybusiness.com",
    "companyincorporationvietnam.com",
    "aurigininc.com",
    "pitchbook.com",
    "alphasearch.com",
    "linkedin.com",
    "itviec.com",
    "vietnamworks.com",
    "topcv.vn",
    "jobstreet.com",
    "wellfound.com",
    "freelancer.com",
    "upwork.com",
    "headhuntvietnam.com",
    "remotepeople.com",
    "gov.vn",
    "chinhphu.vn",
    "eib.org",
    "worldbank.org",
    "adb.org",
    "gov.uk",
    "scb.co.th",
    "scbeic.com",
    "finnomena.com",
    "settrade.com",
    "ceylinco-insurance.com",
    "labuanibfc.com",
    "thaiware.com",
    "uptodown.com",
    "shopify.com",
    "salesforce.com",
    "openai.com",
    "microsoft.com",
    "google.com",
    "amazon.com",
    "apple.com",
    "ibm.com",
    "kpmg.com",
    "ey.com",
    "deloitte.com",
    "pwc.com",
    "lendingpoint.com",
    "transfez.com",
    "gameloft.com",
    "geeksforgeeks.org",
    "w3schools.com",
    "stackoverflow.com",
    "reddit.com",
    "quora.com",
    "poki.com",
    "crazygames.com",
    "bgames.com",
    "hahagames.com",
    "kizi.com",
    "cellphones.com.vn",
    "thegioididong.com",
    "fptshop.com.vn",
    "mobilecity.vn",
    "nguyenkim.com",
    "dienmayxanh.com",
    "topzone.vn",
    "tiki.vn",
    "shopee.vn",
    "lazada.vn",
    "sitestat.com",
    "siteindices.com",
    "usitestat.com",
    "prsync.com",
    "medium.com",
    "substack.com",
    "blogspot.com",
    "wordpress.com",
    "10times.com",
    "eventbrite.com",
    "merriam-webster.com",
    "dictionary.com",
    "wikipedia.org",
    "britannica.com",
    "investopedia.com",
    "wikihow.com",
    "vietjack.com",
    "hoc247.net",
    "tailieu.vn",
    "thivien.net",
    "kahoot.com",
    "runoob.com",
    "buyabans.com",
    "othoba.com",
    "logitech.com",
    "lonelyplanet.com",
    "tripadvisor.com",
    "booking.com",
    "agoda.com",
}

REJECT_DESC_PHRASES = [
    "latest news",
    "breaking news",
    "tin tức mới",
    "read more articles",
    "read our blog",
    "in this article",
    "in this guide",
    "the meaning of",
    "definition of",
    "learn what",
    "lý thuyết",
    "bài tập",
    "học sinh",
    "market size",
    "market share",
    "cagr",
    "market forecast",
    "key players include",
    "table of contents",
    "buy report",
    "download report",
    "find companies",
    "list of companies",
    "top companies in",
    "best companies in",
    "danh sách công ty",
    "bitcoin block explorer",
    "nft marketplace",
    "mutual fund",
    "fund performance",
    "nav history",
    "personal loan calculator",
    "online games",
    "play free",
    "browser games",
    "press release newswire",
    "scholarship",
    "tuition",
    "admissions",
    "road passenger authority",
    "mục lục",
    "fintech là gì",
    "1. fintech",
]

BAD_EMAIL_PATTERNS = [
    r"sentry",
    r"wixpress",
    r"bug-report",
    r"^name@",
    r"^your@",
    r"^youname@",
    r"^enteryour@",
    r"@email\.com$",
    r"@yourcompany\.com$",
    r"@example\.com$",
    r"^test@",
    r"^demo@",
    r"^u003e",
    r"photo-shared-by",
    r"@company\.com$",
    r"@mail\.com$",
    r"^[a-f0-9]{20,}@",
    r"u002f@",
    r"@avif$",
    r"@2x\.",
    r"@robots\.net$",
    r"@benzingaheadlines",
    r"@sharecomms",
    r"^nckh\.",
    r"^efilingwebmanager",
    r"@incometax",
    r"^--",
    r"@tradepassglobal",
    r"@asiapevc",
    r"@eventsnewsasia",
    r"@dealstreetasia",
    r"@businessnewsasia",
    r"@lhdfirm\.com$",
    r"bridgewest\.developer@",
    r"musarurwaregis@gmail",
    r"vietnamoutsourcinghub@gmail",
    r"finsharevi@gmail",
    r"@sohu\.com$",
    r"@qq\.com$",
    r"@163\.com$",
    r"@126\.com$",
    r"@sina\.com$",
    r"^\d{5,}@",
    r"^user@",
    r"^example@",
    r"@myorg\.com$",
    r"_@",
]

BAD_EMAIL_DOMAINS_EXTRA = {
    "sentry.io",
    "wixpress.com",
    "example.com",
    "yourcompany.com",
    "robots.net",
    "benzingaheadlines.com",
    "sharecomms.co.uk",
    "tradepassglobal.com",
    "peimedia.com",
    "asiapevc.com",
    "eventsnewsasia.com",
    "dealstreetasia.com",
    "businessnewsasia.com",
    "soa.org",
    "lhdfirm.com",
    "schema.org",
    "w3.org",
    "cloudflare.com",
    "sohu.com",
    "qq.com",
    "163.com",
    "126.com",
    "sina.com",
    "myorg.com",
    "trangvangvietnam.com",
}

FOREIGN_EMAIL_TLDS = {".fi", ".bd", ".lk", ".pk", ".th", ".cn"}


# ══════════════════════════════════════════════════════════════════════
#  FILTER HELPERS
# ══════════════════════════════════════════════════════════════════════
def _has_value(val):
    return bool(
        val and str(val).strip() not in ("", "nan", "None", "NaN", "false", "False")
    )


def _get_domain(url):
    try:
        from urllib.parse import urlparse

        return urlparse(url).netloc.lower().replace("www.", "")
    except:
        return ""


def _root_domain(url):
    try:
        parts = str(url).split("//")[-1].split("/")[0].split(".")
        return ".".join(parts[-2:]) if len(parts) >= 2 else str(url)
    except:
        return str(url)


def _is_noise_domain(url):
    if not url:
        return True
    dom = _get_domain(url)
    if dom in NOISE_DOMAINS:
        return True
    for nd in NOISE_DOMAINS:
        if dom.endswith("." + nd):
            return True
    return False


def _desc_ok(desc):
    if not desc or str(desc).strip() in ("", "nan", "None"):
        return True
    low = desc.lower()
    return not any(p in low for p in REJECT_DESC_PHRASES)


def _is_bad_email(email):
    if not email or "@" not in email or len(email) > 100:
        return True
    low = email.lower()
    domain = low.split("@")[-1]
    prefix = low.split("@")[0]
    if domain in BAD_EMAIL_DOMAINS_EXTRA:
        return True
    for tld in FOREIGN_EMAIL_TLDS:
        if domain.endswith(tld):
            return True
    if any(re.search(p, low) for p in BAD_EMAIL_PATTERNS):
        return True
    if len(prefix) > 50:
        return True
    return False


def _email_type(email):
    if not email or "@" not in email:
        return "none"
    dom = email.split("@")[-1].lower().strip()
    return "personal" if dom in PERSONAL_DOMAINS else "company"


def _reject_reason(d, cfg):
    if _is_noise_domain(str(d.get("website", ""))):
        return "noise_domain"
    if not _desc_ok(str(d.get("description", ""))):
        return "bad_description"
    grade = str(d.get("grade", "")).strip().upper()
    if grade not in cfg["grades"]:
        return f"grade_{grade or 'missing'}"
    try:
        score = float(d.get("score", 0) or 0)
    except:
        score = 0
    if score < cfg["min_score"]:
        return f"score_{int(score)}"
    website = str(d.get("website", "")).strip()
    if not website or not website.startswith("http"):
        return "no_website"
    if cfg["require_email"]:
        em = str(d.get("best_email", "") or d.get("emails", "") or "").strip()
        first = em.split(",")[0].strip() if em else ""
        if not first or _is_bad_email(first):
            return "bad_email"
    if cfg["require_phone"]:
        if not _has_value(d.get("phones", "")):
            return "no_phone"
    if cfg["industries"]:
        ind = str(d.get("industry", "")).lower()
        if ind not in [i.lower() for i in cfg["industries"]]:
            return f"industry_{ind or '?'}"
    return ""


def apply_filter(df, cfg):
    passed, rejected = [], []
    seen_emails = set()
    seen_domains = set()
    for _, row in df.iterrows():
        d = row.to_dict()
        reason = _reject_reason(d, cfg)
        if reason:
            rejected.append({**d, "reject_reason": reason})
            continue
        # Dedup
        em = (
            str(d.get("best_email", "") or d.get("emails", "") or "")
            .split(",")[0]
            .strip()
            .lower()
        )
        dom = _root_domain(str(d.get("website", "")))
        if em and em not in ("", "nan", "none"):
            if em in seen_emails:
                rejected.append({**d, "reject_reason": "duplicate_email"})
                continue
            seen_emails.add(em)
        if dom:
            if dom in seen_domains:
                rejected.append({**d, "reject_reason": "duplicate_domain"})
                continue
            seen_domains.add(dom)
        passed.append(d)

    df_pass = pd.DataFrame(passed) if passed else pd.DataFrame()
    df_rej = pd.DataFrame(rejected) if rejected else pd.DataFrame()
    if not df_pass.empty:

        def _fe(r):
            em = str(r.get("best_email", "") or r.get("emails", "") or "")
            return em.split(",")[0].strip() if em else ""

        df_pass["email_type"] = df_pass.apply(
            lambda r: _email_type(_fe(r.to_dict())), axis=1
        )
        df_pass["has_phone"] = df_pass.get(
            "phones", pd.Series([""] * len(df_pass))
        ).apply(
            lambda p: bool(str(p).strip() and str(p).strip() not in ("nan", "None", ""))
        )
        go = {"A": 0, "B": 1, "C": 2, "D": 3}
        df_pass["_g"] = df_pass["grade"].apply(lambda g: go.get(str(g).upper(), 3))
        df_pass["_s"] = pd.to_numeric(df_pass.get("score", 0), errors="coerce").fillna(
            0
        )
        df_pass["_et"] = df_pass["email_type"].map(
            {"company": 0, "personal": 1, "none": 2, "unknown": 1}
        )
        df_pass.sort_values(
            ["_g", "_et", "_s"], ascending=[True, True, False], inplace=True
        )
        df_pass.drop(columns=["_g", "_s", "_et"], inplace=True, errors="ignore")
        df_pass.reset_index(drop=True, inplace=True)
    return df_pass, df_rej


# ══════════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════
GRADE_COLORS = {"A": "C6EFCE", "B": "FFEB9C", "C": "DDEEFF", "D": "FFCCCC"}
OUTPUT_COLS = [
    "grade",
    "score",
    "website",
    "best_email",
    "email_type",
    "email_quality",
    "emails",
    "phones",
    "has_phone",
    "field",
    "industry",
    "entity_type",
    "has_contact_page",
    "has_services_page",
    "has_about_page",
    "is_hiring",
    "linkedin_url",
    "description",
    "pain_point",
    "company_type",
    "contact_reason",
    "email_subject",
    "email_body",
    "ai_language",
    "tags",
    "issues",
    "session",
    "should_contact",
]
COL_WIDTHS = {
    "grade": 7,
    "score": 7,
    "website": 42,
    "best_email": 32,
    "email_type": 11,
    "email_quality": 13,
    "emails": 35,
    "phones": 20,
    "has_phone": 9,
    "field": 16,
    "industry": 12,
    "entity_type": 14,
    "has_contact_page": 13,
    "has_services_page": 14,
    "has_about_page": 12,
    "is_hiring": 9,
    "linkedin_url": 35,
    "description": 55,
    "pain_point": 35,
    "company_type": 16,
    "contact_reason": 30,
    "email_subject": 40,
    "email_body": 60,
    "ai_language": 10,
    "tags": 22,
    "issues": 20,
    "session": 18,
    "should_contact": 13,
    "reject_reason": 24,
}
WRAP_COLS = {
    "description",
    "email_body",
    "email_subject",
    "pain_point",
    "contact_reason",
    "tags",
    "issues",
}
CTR_COLS = {
    "grade",
    "score",
    "has_phone",
    "has_contact_page",
    "has_services_page",
    "has_about_page",
    "is_hiring",
    "should_contact",
    "email_type",
    "industry",
    "field",
    "ai_language",
}


def _write_excel(df_pass, df_rej, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HFILL = PatternFill("solid", fgColor="1F4E79")
    HFONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    DFONT = Font(name="Calibri", size=9)
    THIN = Side(style="thin", color="CCCCCC")
    BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR_A = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LFT_A = Alignment(horizontal="left", vertical="top", wrap_text=True)
    LFN_A = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def write_sheet(ws, df, extra_cols=None):
        if df is None or df.empty:
            ws["A1"] = "Không có dữ liệu."
            return
        cols = [c for c in (OUTPUT_COLS + (extra_cols or [])) if c in df.columns]
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(1, ci, col)
            cell.font, cell.fill, cell.alignment, cell.border = HFONT, HFILL, CTR_A, BDR
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 16)
        ws.row_dimensions[1].height = 22
        grade_ci = cols.index("grade") + 1 if "grade" in cols else None
        for ri, row in enumerate(df.itertuples(index=False), 2):
            grade = str(getattr(row, "grade", "")).upper().strip() if grade_ci else ""
            rfill = PatternFill("solid", fgColor=GRADE_COLORS.get(grade, "FFFFFF"))
            for ci, col in enumerate(cols, 1):
                val = getattr(row, col, "") if col in df.columns else ""
                if pd.isna(val):
                    val = ""
                cell = ws.cell(ri, ci, str(val) if val != "" else "")
                cell.font, cell.fill, cell.border = DFONT, rfill, BDR
                cell.alignment = (
                    LFT_A if col in WRAP_COLS else (CTR_A if col in CTR_COLS else LFN_A)
                )
            ws.row_dimensions[ri].height = 14
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "✅ Leads"
    write_sheet(ws1, df_pass)
    ws2 = wb.create_sheet("❌ Rejected")
    write_sheet(ws2, df_rej, extra_cols=["reject_reason"])
    ws3 = wb.create_sheet("📊 Stats")
    hf = Font(bold=True, name="Calibri", size=10)
    bf = Font(name="Calibri", size=9)
    h1 = Font(bold=True, size=11, name="Calibri")

    def section(title, rows, r):
        ws3.cell(r, 1, title).font = h1
        r += 1
        for k, v in rows:
            ws3.cell(r, 1, str(k)).font = hf
            ws3.cell(r, 2, str(v)).font = bf
            r += 1
        return r + 1

    total = len(df_pass) + (len(df_rej) if df_rej is not None else 0)
    r = 1
    r = section(
        "📊 Summary",
        [
            ("Total", total),
            ("✅ Leads", len(df_pass)),
            ("❌ Rejected", len(df_rej) if df_rej is not None else 0),
            ("Pass rate", f"{len(df_pass)/total*100:.1f}%" if total else "0%"),
            (
                "Cold emails",
                (
                    df_pass["email_body"].apply(lambda x: bool(str(x).strip())).sum()
                    if not df_pass.empty and "email_body" in df_pass.columns
                    else 0
                ),
            ),
        ],
        r,
    )
    if not df_pass.empty:
        if "grade" in df_pass.columns:
            r = section(
                "🏆 Grade", list(df_pass["grade"].value_counts().to_dict().items()), r
            )
        if "email_type" in df_pass.columns:
            r = section(
                "📧 Email Type",
                list(df_pass["email_type"].value_counts().to_dict().items()),
                r,
            )
        if "field" in df_pass.columns:
            r = section(
                "🏭 Top Fields",
                list(df_pass["field"].value_counts().head(8).to_dict().items()),
                r,
            )
    if df_rej is not None and not df_rej.empty and "reject_reason" in df_rej.columns:
        r = section(
            "❌ Reject Reasons",
            list(df_rej["reject_reason"].value_counts().head(10).to_dict().items()),
            r,
        )
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 20
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════
#  CRAWLER JOB — dùng logic từ crawler.py (YellowPages)
# ══════════════════════════════════════════════════════════════════════
def run_crawl_job(params):
    import sys, requests, urllib3
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from bs4 import BeautifulSoup

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

    from extractor import extract_contact_and_field
    from filter import filter_url, clean_emails, clean_phones, validate_description
    from scorer import score_lead
    from exporter import export_to_excel

    industry = params["industry"]
    target = int(params.get("target", 100))
    max_workers = int(params.get("workers", 6))
    skip_enrich = params.get("skip_enrich", False)
    skip_ai = params.get("skip_ai", True)
    session = datetime.now().strftime("%Y%m%d_%H%M%S")

    UA = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    ]

    def hdrs():
        return {
            "User-Agent": random.choice(UA),
            "Accept": "text/html,*/*;q=0.8",
            "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.8",
            "Referer": "https://www.yellowpages.vn/",
        }

    def clog(msg):
        with _lock:
            _crawl["logs"].append(str(msg))

    def set_phase(p):
        with _lock:
            _crawl["phase"] = p

    def get_status():
        with _lock:
            return _crawl["status"]

    def _get_dom(url):
        try:
            return url.split("//")[-1].split("/")[0].lower().replace("www.", "")
        except:
            return ""

    def fetch(url, timeout=12):
        for attempt in range(2):
            try:
                r = requests.get(
                    url,
                    headers=hdrs(),
                    timeout=timeout,
                    verify=False,
                    allow_redirects=True,
                )
                if r.status_code == 200:
                    return r
                if r.status_code in (429, 503):
                    time.sleep(5 * (attempt + 1))
            except:
                if attempt == 0:
                    time.sleep(2)
        return None

    # ── Phase 1: Lấy /lgs/ links từ YellowPages ──────────────
    try:
        clog(f"{'='*55}")
        clog(f"🎯 {industry.upper()} | target={target} | workers={max_workers}")
        clog(f"📅 Session: {session}")
        clog(f"{'='*55}\n")
        set_phase("Lấy danh sách công ty từ YellowPages")
        clog("📋 PHASE 1: Lấy danh sách công ty từ YellowPages\n")

        cats = YP_CATEGORIES.get(industry, [])
        if not cats:
            clog(f"❌ Không có category cho ngành: {industry}")
            with _lock:
                _crawl["status"] = "error"
                return

        lgs_links = []
        seen_lgs = set()

        for cat_url in cats:
            if get_status() == "stopping":
                break
            slug = cat_url.split("/")[-1][:50]
            clog(f"  📂 {slug}")
            for page in range(1, 6):
                if get_status() == "stopping":
                    break
                url = (
                    cat_url
                    if page == 1
                    else (
                        cat_url + f"&page={page}"
                        if "?" in cat_url
                        else cat_url.replace(".html", f"-p{page}.html")
                    )
                )
                r = fetch(url)
                if not r:
                    break
                soup = BeautifulSoup(r.text, "html.parser")
                found = 0
                for a in soup.select("a[href]"):
                    href = a.get("href", "")
                    if "/lgs/" in href:
                        full = (
                            "https://www.yellowpages.vn" + href
                            if href.startswith("/")
                            else href
                        )
                        if full not in seen_lgs:
                            seen_lgs.add(full)
                            lgs_links.append((full, industry))
                            found += 1
                clog(f"    trang {page}: +{found} ({len(lgs_links)} total)")
                if found == 0:
                    break
                time.sleep(random.uniform(0.8, 1.5))
            if len(lgs_links) >= target * 3:
                break

        clog(f"\n✅ {len(lgs_links)} công ty trong danh sách\n")
        if not lgs_links:
            clog("⚠️  Không lấy được danh sách công ty.")
            with _lock:
                _crawl["status"] = "done"
                return

        # ── Phase 2: Lấy website từ mỗi trang /lgs/ ──────────
        set_phase("Lấy website từ từng công ty")
        clog("🔗 PHASE 2: Lấy website thật từ trang công ty\n")

        website_queue = []
        seen_domains = set()

        for i, (lgs_url, ind) in enumerate(lgs_links):
            if get_status() == "stopping":
                break
            r = fetch(lgs_url)
            if not r:
                clog(f"  ❌ [{i+1}] no resp: {lgs_url[-40:]}")
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            company_name = ""
            h1 = soup.find("h1")
            if h1:
                company_name = h1.get_text(strip=True)
            website = ""
            for a in soup.select("a[href]"):
                href = a.get("href", "")
                if not href.startswith("http"):
                    continue
                dom = _get_dom(href)
                if dom and dom not in SKIP_DOMAINS_YP and len(dom) > 4:
                    website = href.split("?")[0].rstrip("/")
                    break
            if not website:
                text = soup.get_text()
                m = re.search(
                    r"(?:www\.)?([a-z0-9\-]+\.(?:com\.vn|vn|net|org))", text, re.I
                )
                if m:
                    website = "https://" + m.group(0).lower()
            if website:
                dom = _get_dom(website)
                if dom and dom not in seen_domains and dom not in SKIP_DOMAINS_YP:
                    seen_domains.add(dom)
                    website_queue.append((website, ind, company_name))
                    clog(
                        f"  ✅ [{i+1}/{len(lgs_links)}] {company_name[:30]} → {website[:50]}"
                    )
                else:
                    clog(f"  ⏭  [{i+1}/{len(lgs_links)}] dup: {website[:50]}")
            else:
                clog(f"  ❌ [{i+1}/{len(lgs_links)}] no website: {lgs_url[-40:]}")
            time.sleep(random.uniform(0.4, 0.8))
            if len(website_queue) >= target * 2:
                clog(f"  → Đủ {len(website_queue)} websites")
                break

        clog(f"\n✅ {len(website_queue)} websites để crawl\n")
        with _lock:
            _crawl["total_urls"] = len(website_queue)
            _crawl["done_urls"] = 0

        if not website_queue:
            clog("⚠️  Không lấy được website nào.")
            with _lock:
                _crawl["status"] = "done"
                return

        # ── Phase 3: Crawl từng website ───────────────────────
        set_phase("Crawl website — extract email/phone")
        clog(
            f"🕷️  PHASE 3: CRAWL {len(website_queue)} websites ({max_workers} threads)\n"
        )

        raw_leads = []
        done = [0]
        done_lock = threading.Lock()

        def _meta_desc(html):
            try:
                soup = BeautifulSoup(html, "html.parser")
                for attr in [{"name": "description"}, {"property": "og:description"}]:
                    tag = soup.find("meta", attrs=attr)
                    if tag and tag.get("content"):
                        return tag["content"][:300].strip()
                for p in soup.find_all("p"):
                    t = p.get_text(strip=True)
                    if len(t) > 60:
                        return t[:300]
            except:
                pass
            return ""

        def crawl_one(item):
            ws, ind, name = item
            try:
                dom = _get_dom(ws)
                if not dom or dom in SKIP_DOMAINS_YP:
                    return None
                keep, _ = filter_url(ws)
                if not keep:
                    return None
                r = fetch(ws, timeout=12)
                if not r:
                    return None
                html = r.text
                desc = _meta_desc(html)
                # Fix encoding
                try:
                    desc = desc.encode("latin-1").decode("utf-8")
                except:
                    pass
                if desc.count("á»") > 2 or desc.count("Ã") > 2:
                    desc = ""
                keep, _ = validate_description(desc)
                if not keep:
                    return None
                emails_raw, phones_raw, field = extract_contact_and_field(ws)
                from filter import clean_emails as ce, clean_phones as cp

                emails = ce(emails_raw)
                phones = cp(phones_raw)
                if not emails and not phones:
                    return None
                # Reject trangvangvietnam placeholder emails
                if emails and all(
                    "trangvangvietnam" in e
                    or e in ["email@congty.vn", "info@congty.vn", "email@domain.com"]
                    for e in emails
                ):
                    return None
                # VN check: .vn domain HOẶC có phone VN
                is_vn = dom.endswith(".vn") or ".vn." in dom
                has_vn_phone = bool(phones)
                if not is_vn and not has_vn_phone:
                    return None
                return {
                    "website": ws,
                    "company_name": name,
                    "emails": ", ".join(emails[:5]),
                    "phones": ", ".join(phones[:3]),
                    "field": field,
                    "description": desc,
                    "industry": ind,
                    "session": session,
                }
            except:
                return None

        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            fmap = {pool.submit(crawl_one, item): item for item in website_queue}
            for fut in as_completed(fmap):
                with done_lock:
                    done[0] += 1
                    with _lock:
                        _crawl["done_urls"] = done[0]
                if get_status() == "stopping":
                    pool.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    result = fut.result()
                except:
                    result = None
                bar = f"[{done[0]}/{len(website_queue)}|{len(raw_leads)}↑]"
                if result:
                    raw_leads.append(result)
                    clog(
                        f"  ✅ {bar} {result['field']} | {result.get('emails','')[:30]} | {result['website'][:45]}"
                    )
                else:
                    item = fmap[fut]
                    clog(f"  ❌ {bar} {item[0][:60]}")

        clog(f"\n📊 Raw leads có contact: {len(raw_leads)}\n")
        if not raw_leads:
            clog("⚠️  Không có lead nào có email/phone.")
            with _lock:
                _crawl["status"] = "done"
                return

        # ── Phase 4: Enrich ───────────────────────────────────
        if not skip_enrich:
            set_phase("Enrich — kiểm tra trang web")
            clog(f"🔍 PHASE 4: ENRICH ({len(raw_leads)} leads)\n")
            from enricher import enrich_all

            raw_leads = enrich_all(raw_leads, max_workers=6)

        # ── Phase 5: Score ────────────────────────────────────
        set_phase("Score & Grade leads")
        clog("📊 PHASE 5: SCORE\n")
        scored = [score_lead(l) for l in raw_leads]
        before = len(scored)
        scored = [l for l in scored if l.get("grade", "D") != "D"]
        scored.sort(
            key=lambda x: (
                {"A": 0, "B": 1, "C": 2, "D": 3}.get(x.get("grade", "D"), 3),
                -x.get("score", 0),
            )
        )
        gc = {}
        for l in scored:
            g = l.get("grade", "?")
            gc[g] = gc.get(g, 0) + 1
        clog(f"  → {len(scored)} leads (removed {before-len(scored)} Grade D) | {gc}")

        # ── Phase 6: AI (optional) ────────────────────────────
        if not skip_ai and scored:
            try:
                from ai_analyst import analyze_leads

                clog(f"\n🤖 PHASE 6: AI ANALYZE")
                set_phase("AI analyze")
                scored = analyze_leads(scored, industry)
            except Exception as e:
                clog(f"⚠️  AI skip: {e}")

        # ── Phase 7: Export ───────────────────────────────────
        set_phase("Export Excel")
        clog(f"\n📤 EXPORT → leads_{industry}_{session}.xlsx")
        export_to_excel(scored, session=session, industry=industry)

        clog(f"\n{'='*55}")
        clog(f"🎉 DONE | {len(scored)} leads | {industry.upper()} | {session}")
        for g, c in sorted(gc.items()):
            clog(f"   Grade {g}: {c}")
        clog(f"{'='*55}")

        with _lock:
            _crawl["leads"] = scored
            _crawl["status"] = "done"
            _crawl["session"] = session
            try:
                sessions = []
                if os.path.exists("sessions.json"):
                    sessions = json.load(open("sessions.json", "r", encoding="utf-8"))
                sessions.insert(
                    0, {"industry": industry, "leads": len(scored), "session": session}
                )
                json.dump(
                    sessions[:20],
                    open("sessions.json", "w", encoding="utf-8"),
                    ensure_ascii=False,
                )
            except:
                pass

    except Exception as e:
        import traceback

        clog(f"\n❌ Lỗi: {e}")
        clog(traceback.format_exc())
        with _lock:
            _crawl["status"] = "error"


# ══════════════════════════════════════════════════════════════════════
#  HTML TEMPLATE
# ══════════════════════════════════════════════════════════════════════
HTML = r"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Lead.AI — Lead Intelligence Tool</title>
<style>
:root{--bg:#0b0d15;--surface:#12151f;--border:#1e2235;--accent:#3b82f6;--green:#22c55e;--yellow:#eab308;--red:#ef4444;--purple:#a855f7;--text:#e2e8f0;--muted:#64748b;--card:#161929}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--text);height:100vh;display:flex;flex-direction:column;overflow:hidden}
header{background:var(--surface);border-bottom:1px solid var(--border);padding:0 24px;height:52px;display:flex;align-items:center;gap:14px;flex-shrink:0}
.logo{font-size:1.2rem;font-weight:800;letter-spacing:-.5px;color:#fff}.logo span{color:var(--accent)}
.ver{font-size:.7rem;background:#1e2235;color:var(--muted);padding:2px 8px;border-radius:20px}
.nav-tabs{display:flex;margin-left:20px;height:100%}
.nav-tab{display:flex;align-items:center;gap:7px;padding:0 20px;font-size:.85rem;color:var(--muted);cursor:pointer;border-bottom:2px solid transparent;transition:all .15s;white-space:nowrap}
.nav-tab:hover{color:var(--text)}.nav-tab.active{color:#fff;border-bottom-color:var(--accent)}
.cnt{background:#1e2235;padding:1px 7px;border-radius:10px;font-size:.68rem;margin-left:2px}
.nav-tab.active .cnt{background:var(--accent);color:#fff}
.ml{margin-left:auto;display:flex;gap:8px;align-items:center}
.workspace{display:flex;flex:1;overflow:hidden}
.sidebar{width:280px;background:var(--surface);border-right:1px solid var(--border);overflow-y:auto;padding:16px;flex-shrink:0}
.main{flex:1;overflow:hidden;display:flex;flex-direction:column}
.s-title{font-size:.65rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:1.2px;margin:16px 0 8px}.s-title:first-child{margin-top:0}
.ind-card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:11px 13px;margin-bottom:6px;cursor:pointer;transition:all .15s;position:relative;overflow:hidden}
.ind-card:hover{border-color:#2d3555}.ind-card.sel{border-color:var(--ac,var(--accent));background:#131828}
.ind-card.sel::before{content:'';position:absolute;left:0;top:0;bottom:0;width:3px;background:var(--ac,var(--accent))}
.ind-top{display:flex;align-items:center;justify-content:space-between;gap:8px}
.ind-label{font-size:.88rem;font-weight:600}.ind-badge{font-size:.62rem;padding:2px 7px;border-radius:20px;background:var(--ac,var(--accent));color:#fff;opacity:.75}
.ind-desc{font-size:.75rem;color:var(--muted);margin-top:4px}
.ind-why{font-size:.72rem;color:var(--muted);margin-top:6px;padding:5px 8px;background:#0d1018;border-radius:5px;display:none;border-left:2px solid var(--ac,var(--accent))}
.ind-card.sel .ind-why{display:block}
.field-row{display:flex;align-items:center;gap:8px;margin-bottom:8px}
.field-row label{font-size:.75rem;color:var(--muted);width:75px;flex-shrink:0}
input[type=number],input[type=text],select{background:#0f1120;border:1px solid var(--border);color:var(--text);padding:7px 10px;border-radius:6px;width:100%;font-size:.82rem}
input:focus,select:focus{outline:none;border-color:var(--accent)}
.tog-row{display:flex;gap:6px;flex-wrap:wrap}
.tog{font-size:.72rem;padding:4px 11px;border-radius:5px;cursor:pointer;border:1px solid var(--border);background:#0f1120;color:var(--muted);transition:all .15s;user-select:none}
.tog.on{background:#132234;border-color:#2563eb;color:#60a5fa}
.btn{width:100%;padding:11px;border-radius:7px;border:none;cursor:pointer;font-size:.88rem;font-weight:700;transition:all .15s;margin-top:8px}
.btn-blue{background:linear-gradient(135deg,#2563eb,#1d4ed8);color:#fff}
.btn-red{background:linear-gradient(135deg,#dc2626,#b91c1c);color:#fff;display:none}
.btn-green{background:linear-gradient(135deg,#16a34a,#15803d);color:#fff}
.btn-purple{background:linear-gradient(135deg,#7c3aed,#6d28d9);color:#fff}
.btn:hover:not(:disabled){opacity:.88;transform:translateY(-1px)}
.btn:disabled{opacity:.4;cursor:not-allowed;transform:none}.btn.show{display:block}
/* phase badge */
.phase-badge{font-size:.7rem;color:var(--accent);background:#131e36;border:1px solid #1e3a6a;border-radius:5px;padding:3px 9px;margin-top:6px;display:none;text-align:center}
.phase-badge.show{display:block}
/* progress */
.prog-wrap{padding:10px 20px;background:var(--surface);border-bottom:1px solid var(--border);display:none}
.prog-bar{height:3px;background:var(--border);border-radius:2px;overflow:hidden}
.prog-fill{height:100%;background:linear-gradient(90deg,var(--accent),var(--green));border-radius:2px;transition:width .3s}
.prog-txt{font-size:.7rem;color:var(--muted);margin-top:4px;text-align:right}
/* log */
.log-panel{padding:14px 18px;font-family:'Cascadia Code','Consolas',monospace;font-size:.76rem;line-height:1.7;overflow-y:auto;flex:1}
.l{margin-bottom:1px}.l.ok{color:var(--green)}.l.err{color:var(--red)}.l.info{color:var(--accent)}.l.warn{color:var(--yellow)}.l.dim{color:#374151}
/* leads */
.leads-wrap{padding:16px;overflow-y:auto;flex:1}
.stats-row{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}
.stat{background:var(--card);border:1px solid var(--border);border-radius:7px;padding:8px 14px;text-align:center;min-width:72px}
.stat .n{font-size:1.3rem;font-weight:700}.stat .l{font-size:.65rem;color:var(--muted);margin-top:2px}
.stat.A .n{color:var(--yellow)}.stat.B .n{color:var(--green)}.stat.C .n{color:var(--accent)}
.tbl{width:100%;border-collapse:collapse;font-size:.78rem}
.tbl th{background:#0f1120;padding:7px 10px;text-align:left;font-size:.64rem;color:var(--muted);text-transform:uppercase;border-bottom:1px solid var(--border);white-space:nowrap;position:sticky;top:0}
.tbl td{padding:8px 10px;border-bottom:1px solid #131626;vertical-align:middle;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.tbl tr:hover td{background:#131626}
.gp{display:inline-flex;width:22px;height:22px;border-radius:4px;align-items:center;justify-content:center;font-weight:700;font-size:.72rem}
.gA{background:#eab30822;color:var(--yellow)}.gB{background:#22c55e22;color:var(--green)}.gC{background:#3b82f622;color:var(--accent)}.gD{background:#ef444422;color:var(--red)}
.em-dm{color:var(--yellow);font-size:.7rem}.em-s{color:var(--green);font-size:.7rem}.em-g{color:var(--accent);font-size:.7rem}.em-l{color:var(--muted);font-size:.7rem}
a.lnk{color:var(--accent);text-decoration:none}a.lnk:hover{text-decoration:underline}
/* email cards */
.email-card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:14px;margin-bottom:10px}
.ec-site{font-size:.78rem;color:var(--accent);font-weight:600;margin-bottom:3px}
.ec-subj{font-size:.88rem;font-weight:700;color:#fff;margin-bottom:7px}
.ec-body{font-size:.78rem;color:#94a3b8;line-height:1.65;white-space:pre-wrap}
.ec-pain{font-size:.72rem;color:var(--muted);margin-top:6px;font-style:italic}
.ec-meta{font-size:.68rem;color:#374151;margin-top:5px}
/* rejected */
.rej-stat{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px}
.rej-pill{background:#1c0f0f;border:1px solid #3f1515;border-radius:6px;padding:5px 11px;font-size:.72rem;color:var(--red)}
.rej-pill .rn{font-weight:700;font-size:.95rem}
/* upload zone */
.upload-zone{border:1.5px dashed var(--border);border-radius:7px;padding:16px;text-align:center;cursor:pointer;margin-bottom:10px;transition:border-color .15s}
.upload-zone:hover{border-color:var(--accent)}
/* empty */
.empty{text-align:center;padding:60px 20px;color:var(--muted)}.empty .ei{font-size:2.5rem;margin-bottom:10px}
</style>
</head>
<body>
<header>
  <div class="logo">Lead<span>.</span>AI</div>
  <div class="ver">v5 · YellowPages</div>
  <nav class="nav-tabs">
    <div class="nav-tab active" id="nt-crawl" onclick="goTab('crawl')">🕷️ Crawl <span class="cnt" id="cnt-crawl">0</span></div>
    <div class="nav-tab" id="nt-filter" onclick="goTab('filter')">🔍 Filter <span class="cnt" id="cnt-filter">0</span></div>
    <div class="nav-tab" id="nt-email" onclick="goTab('email')">📧 Email <span class="cnt" id="cnt-email">0</span></div>
    <div class="nav-tab" id="nt-rejected" onclick="goTab('rejected')">❌ Rejected <span class="cnt" id="cnt-rejected">0</span></div>
  </nav>
  <div class="ml">
    <button class="btn btn-green" style="width:auto;margin:0;padding:7px 14px;font-size:.78rem"
            onclick="exportExcel()" id="btn-export" disabled>⬇ Export Excel</button>
  </div>
</header>

<div class="workspace">
<!-- SIDEBAR -->
<div class="sidebar">

  <!-- sb-crawl -->
  <div id="sb-crawl">
    <div class="s-title">Chọn ngành</div>
    <div id="ind-list"></div>
    <div class="s-title">Cấu hình</div>
    <div class="field-row"><label>Target</label><input type="number" id="cfg-target" value="60" min="10" max="500"></div>
    <div class="field-row"><label>Workers</label><input type="number" id="cfg-workers" value="6" min="1" max="12"></div>
    <div class="s-title">Options</div>
    <div class="tog-row">
      <div class="tog on" id="tog-enrich" onclick="this.classList.toggle('on')">🔍 Enrich</div>
      <div class="tog" id="tog-ai" onclick="this.classList.toggle('on')">🤖 AI</div>
    </div>
    <button class="btn btn-blue" id="btn-start" onclick="startCrawl()">▶ Bắt đầu crawl</button>
    <button class="btn btn-red" id="btn-stop" onclick="stopCrawl()">⏹ Dừng</button>
    <div class="phase-badge" id="phase-badge">—</div>
    <div class="s-title" style="margin-top:18px">Session gần đây</div>
    <div id="sessions" style="font-size:.73rem;color:var(--muted)">—</div>
  </div>

  <!-- sb-filter -->
  <div id="sb-filter" style="display:none">
    <div class="s-title">Nguồn dữ liệu</div>
    <div class="upload-zone" id="drop-zone"
         onclick="document.getElementById('file-in').click()"
         ondragover="event.preventDefault();this.style.borderColor='var(--accent)'"
         ondragleave="this.style.borderColor=''"
         ondrop="handleDrop(event)">
      <div style="font-size:1.6rem;opacity:.4">📂</div>
      <div style="font-size:.75rem;color:var(--muted);margin-top:5px" id="upload-lbl">
        Kéo thả Excel hoặc click chọn file<br><small>hoặc dùng kết quả từ tab Crawl</small>
      </div>
      <input type="file" id="file-in" accept=".xlsx,.xls,.csv" style="display:none" onchange="handleFile(event)">
    </div>
    <div class="s-title">Filter Config</div>
    <div class="field-row"><label>Min Score</label><input type="number" id="f-score" value="15" min="0" max="100"></div>
    <div class="field-row"><label>Grades</label>
      <div class="tog-row">
        <div class="tog on" id="fg-A" onclick="this.classList.toggle('on')">A</div>
        <div class="tog on" id="fg-B" onclick="this.classList.toggle('on')">B</div>
        <div class="tog on" id="fg-C" onclick="this.classList.toggle('on')">C</div>
      </div>
    </div>
    <div class="field-row"><label>Industry</label><input type="text" id="f-ind" placeholder="hospitality, it..."></div>
    <div class="field-row"><label>Options</label>
      <div class="tog-row">
        <div class="tog on" id="f-req-email" onclick="this.classList.toggle('on')">Email ✓</div>
        <div class="tog" id="f-req-phone" onclick="this.classList.toggle('on')">Phone ✓</div>
      </div>
    </div>
    <button class="btn btn-blue" onclick="runFilter()" id="btn-filter">🔍 Lọc Leads</button>
  </div>

  <!-- sb-email -->
  <div id="sb-email" style="display:none">
    <div class="s-title">Gen Cold Email</div>
    <div style="font-size:.75rem;color:var(--muted);margin-bottom:8px" id="email-info">Filter leads trước để gen email.</div>
    <div style="font-size:.72rem;color:var(--yellow);margin-bottom:10px" id="cost-est"></div>
    <button class="btn btn-purple" onclick="runGenEmail()" id="btn-gen" disabled>🤖 Gen AI Email</button>
    <div class="s-title" style="margin-top:14px">Hướng dẫn</div>
    <div style="font-size:.73rem;color:var(--muted);line-height:1.7">
      1. Lọc leads ở tab Filter<br>
      2. Bấm Gen AI Email<br>
      3. Xem email ở tab này<br>
      4. Export Excel → chỉnh sửa<br>
      5. Gửi từ email cá nhân
    </div>
  </div>

</div><!-- /sidebar -->

<!-- MAIN -->
<div class="main">
  <div class="prog-wrap" id="prog-wrap">
    <div class="prog-bar"><div class="prog-fill" id="prog-fill" style="width:0%"></div></div>
    <div class="prog-txt" id="prog-txt">—</div>
  </div>

  <div id="tab-crawl" style="display:flex;flex-direction:column;flex:1;overflow:hidden">
    <div class="log-panel" id="log-panel">
      <div class="l dim">Chọn ngành và nhấn Bắt đầu để bắt đầu crawl từ YellowPages...</div>
    </div>
  </div>

  <div id="tab-filter" style="display:none;flex-direction:column;flex:1;overflow:hidden">
    <div class="leads-wrap" id="filter-panel">
      <div class="empty"><div class="ei">🔍</div>Upload Excel hoặc crawl xong để lọc leads.</div>
    </div>
  </div>

  <div id="tab-email" style="display:none;flex-direction:column;flex:1;overflow:hidden">
    <div class="leads-wrap" id="email-panel">
      <div class="empty"><div class="ei">📧</div>Gen AI email để xem kết quả.</div>
    </div>
  </div>

  <div id="tab-rejected" style="display:none;flex-direction:column;flex:1;overflow:hidden">
    <div class="leads-wrap" id="rejected-panel">
      <div class="empty"><div class="ei">✅</div>Chưa có dữ liệu.</div>
    </div>
  </div>
</div>
</div><!-- /workspace -->

<script>
const IND = {{ industries|tojson }};
let selInd = null, crawlPoll = null, genPoll = null;
let crawlLeads=[], filteredLeads=[], rejectedLeads=[], emailLeads=[];
let logCount = 0;

// ── TABS ──────────────────────────────────────────────────────────────
const TABS = ['crawl','filter','email','rejected'];
function goTab(name) {
  TABS.forEach(t => {
    document.getElementById('nt-'+t).classList.toggle('active', t===name);
    document.getElementById('tab-'+t).style.display = t===name ? 'flex' : 'none';
    const sb = document.getElementById('sb-'+t);
    if(sb) sb.style.display = t===name ? 'block' : 'none';
  });
}

// ── INDUSTRIES ────────────────────────────────────────────────────────
function renderIndustries() {
  const el = document.getElementById('ind-list');
  el.innerHTML = '';
  Object.entries(IND).forEach(([key, m]) => {
    const d = document.createElement('div');
    d.className = 'ind-card'; d.id = 'ic-'+key;
    d.style.setProperty('--ac', m.color);
    d.onclick = () => selIndustry(key);
    d.innerHTML = `<div class="ind-top">
      <span class="ind-label">${m.label}</span>
      <span class="ind-badge">${m.cats}cat</span>
    </div>
    <div class="ind-desc">${m.desc}</div>
    <div class="ind-why">💡 ${m.why}</div>`;
    el.appendChild(d);
  });
}
function selIndustry(key) {
  if(selInd) document.getElementById('ic-'+selInd)?.classList.remove('sel');
  selInd = key;
  const card = document.getElementById('ic-'+key);
  if(card){ card.classList.add('sel'); card.style.setProperty('--ac', IND[key].color); }
}

// ── CRAWL ─────────────────────────────────────────────────────────────
async function startCrawl() {
  if(!selInd){ alert('Chọn ngành trước!'); return; }
  const btn = document.getElementById('btn-start');
  btn.disabled=true; btn.textContent='⏳ Đang chạy...';
  document.getElementById('btn-stop').classList.add('show');
  document.getElementById('log-panel').innerHTML='';
  document.getElementById('prog-wrap').style.display='block';
  document.getElementById('phase-badge').classList.add('show');
  logCount=0;
  await fetch('/api/crawl/start',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({
      industry: selInd,
      target:   +document.getElementById('cfg-target').value,
      workers:  +document.getElementById('cfg-workers').value,
      skip_enrich: !document.getElementById('tog-enrich').classList.contains('on'),
      skip_ai:     !document.getElementById('tog-ai').classList.contains('on'),
    })});
  crawlPoll = setInterval(pollCrawl, 1000);
}

async function stopCrawl() {
  await fetch('/api/crawl/stop',{method:'POST'});
}

async function pollCrawl() {
  try {
    const res = await fetch('/api/crawl/status');
    const d   = await res.json();
    appendLogs(d.new_logs||[], 'log-panel');
    if(d.phase) document.getElementById('phase-badge').textContent = '⚙️ '+d.phase;
    if(d.total_urls > 0){
      const pct = Math.round(d.done_urls/d.total_urls*100);
      document.getElementById('prog-fill').style.width=pct+'%';
      document.getElementById('prog-txt').textContent=`${d.done_urls}/${d.total_urls} URLs | ${d.leads?.length||0} leads`;
    }
    if(d.leads?.length){ crawlLeads=d.leads; renderLeadsTable(crawlLeads, 'filter-panel'); }
    if(d.status !== 'running'){
      clearInterval(crawlPoll);
      document.getElementById('btn-start').disabled=false;
      document.getElementById('btn-start').textContent='▶ Bắt đầu crawl';
      document.getElementById('btn-stop').classList.remove('show');
      document.getElementById('phase-badge').classList.remove('show');
      if(d.leads?.length){
        document.getElementById('btn-export').disabled=false;
        document.getElementById('cnt-crawl').textContent=d.leads.length;
        // Auto-load vào filter
        fetch('/api/filter/load-crawl',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({leads:d.leads})});
        document.getElementById('upload-lbl').textContent=`✅ ${d.leads.length} leads từ crawl — sẵn sàng lọc`;
      }
      loadSessions();
    }
  } catch(e){}
}

function appendLogs(lines, panelId) {
  const panel = document.getElementById(panelId);
  if(!panel) return;
  lines.forEach(line => {
    const div = document.createElement('div');
    div.className='l '+logClass(line); div.textContent=line;
    panel.appendChild(div); logCount++;
  });
  document.getElementById('cnt-crawl').textContent=logCount;
  panel.scrollTop=panel.scrollHeight;
}

function logClass(line){
  if(line.includes('✅')) return 'ok';
  if(line.includes('❌')) return 'err';
  if(line.includes('⚠️')||line.includes('⏳')) return 'warn';
  if(line.match(/[🔍🕷️📊🎯📤🏭🤖📋🔗]/u)) return 'info';
  return '';
}

// ── LEADS TABLE ──────────────────────────────────────────────────────
function renderLeadsTable(leads, panelId='filter-panel') {
  document.getElementById('cnt-filter').textContent=leads.length;
  const gc={A:0,B:0,C:0};
  leads.forEach(l=>{if(gc[l.grade]!==undefined)gc[l.grade]++;});
  let html=`<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
    <div class="stats-row">
      <div class="stat A"><div class="n">${gc.A}</div><div class="l">Grade A</div></div>
      <div class="stat B"><div class="n">${gc.B}</div><div class="l">Grade B</div></div>
      <div class="stat C"><div class="n">${gc.C}</div><div class="l">Grade C</div></div>
      <div class="stat"><div class="n" style="color:var(--accent)">${leads.length}</div><div class="l">Total</div></div>
    </div>
  </div>
  <table class="tbl"><thead><tr>
    <th>Grade</th><th>Website</th><th>Best Email</th><th>Quality</th>
    <th>Phone</th><th>Field</th><th>Score</th>
  </tr></thead><tbody>`;
  leads.forEach(l=>{
    const g=l.grade||'?', eq=l.email_quality||'';
    const ec=eq.includes('Decision')?'em-dm':eq.includes('Sales')?'em-s':eq.includes('Generic')?'em-g':'em-l';
    const site=(()=>{try{return new URL(l.website).hostname.replace('www.','')}catch{return l.website||''}})();
    html+=`<tr>
      <td><span class="gp g${g}">${g}</span> <small style="color:var(--muted)">+${l.score||0}</small></td>
      <td><a class="lnk" href="${l.website}" target="_blank">${site}</a></td>
      <td>${l.best_email||'—'}</td>
      <td><span class="${ec}">${eq||'—'}</span></td>
      <td>${(l.phones||'').split(',')[0]||'—'}</td>
      <td>${l.field||'—'}</td>
      <td>${l.score||0}</td>
    </tr>`;
  });
  html+='</tbody></table>';
  document.getElementById(panelId).innerHTML=html;
}

// ── UPLOAD ────────────────────────────────────────────────────────────
function handleDrop(e){
  e.preventDefault();
  document.getElementById('drop-zone').style.borderColor='';
  if(e.dataTransfer.files[0]) uploadFile(e.dataTransfer.files[0]);
}
function handleFile(e){if(e.target.files[0])uploadFile(e.target.files[0]);}
async function uploadFile(file){
  document.getElementById('upload-lbl').textContent=`⏳ Đọc ${file.name}...`;
  const fd=new FormData(); fd.append('file',file);
  const res=await fetch('/api/filter/upload',{method:'POST',body:fd});
  const d=await res.json();
  document.getElementById('upload-lbl').textContent=d.ok?`✅ ${file.name} (${d.rows} rows)`:`❌ ${d.error}`;
}

// ── FILTER ────────────────────────────────────────────────────────────
async function runFilter(){
  const btn=document.getElementById('btn-filter');
  btn.disabled=true; btn.textContent='⏳ Đang lọc...';
  const cfg={
    min_score:     +document.getElementById('f-score').value||15,
    grades:        ['A','B','C'].filter(g=>document.getElementById('fg-'+g).classList.contains('on')),
    require_email: document.getElementById('f-req-email').classList.contains('on'),
    require_phone: document.getElementById('f-req-phone').classList.contains('on'),
    industries:    document.getElementById('f-ind').value.split(',').map(s=>s.trim()).filter(Boolean),
  };
  try{
    const res=await fetch('/api/filter/run',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)});
    const d=await res.json();
    if(d.ok){
      filteredLeads=d.passed; rejectedLeads=d.rejected;
      renderLeadsTable(filteredLeads,'filter-panel');
      renderRejected(rejectedLeads);
      document.getElementById('cnt-filter').textContent=filteredLeads.length;
      document.getElementById('cnt-rejected').textContent=rejectedLeads.length;
      document.getElementById('btn-export').disabled=false;
      document.getElementById('btn-gen').disabled=filteredLeads.length===0;
      document.getElementById('email-info').textContent=`${filteredLeads.length} leads sẵn sàng gen email.`;
      document.getElementById('cost-est').textContent=d.cost_estimate||'';
      goTab('filter');
    } else{ alert('Lỗi: '+d.error); }
  }catch(e){alert('Filter lỗi: '+e);}
  btn.disabled=false; btn.textContent='🔍 Lọc Leads';
}

// ── REJECTED ─────────────────────────────────────────────────────────
function renderRejected(leads){
  if(!leads.length){
    document.getElementById('rejected-panel').innerHTML='<div class="empty"><div class="ei">✅</div>Không có lead bị loại.</div>';
    return;
  }
  const rc={};
  leads.forEach(l=>{rc[l.reject_reason]=(rc[l.reject_reason]||0)+1;});
  let html=`<div class="rej-stat">${
    Object.entries(rc).sort((a,b)=>b[1]-a[1]).slice(0,8)
    .map(([k,v])=>`<div class="rej-pill"><div class="rn">${v}</div>${k}</div>`).join('')
  }</div>
  <table class="tbl"><thead><tr><th>Grade</th><th>Website</th><th>Lý do</th><th>Field</th><th>Score</th></tr></thead><tbody>`;
  leads.slice(0,150).forEach(l=>{
    const g=l.grade||'?';
    const site=(()=>{try{return new URL(l.website).hostname.replace('www.','')}catch{return l.website||''}})();
    html+=`<tr>
      <td><span class="gp g${g}">${g}</span></td>
      <td><a class="lnk" href="${l.website}" target="_blank">${site}</a></td>
      <td style="color:var(--red)">${l.reject_reason||'?'}</td>
      <td>${l.field||'—'}</td><td>${l.score||0}</td>
    </tr>`;
  });
  html+='</tbody></table>';
  document.getElementById('rejected-panel').innerHTML=html;
}

// ── GEN EMAIL ────────────────────────────────────────────────────────
async function runGenEmail(){
  const btn=document.getElementById('btn-gen');
  btn.disabled=true; btn.textContent='⏳ Đang gen...';
  document.getElementById('prog-wrap').style.display='block';
  document.getElementById('email-panel').innerHTML='<div class="log-panel" id="gen-log" style="height:100%"><div class="l info">🤖 Bắt đầu gen email...</div></div>';
  goTab('email');
  await fetch('/api/email/gen',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({leads:filteredLeads})});
  genPoll=setInterval(pollGen,1000);
}
async function pollGen(){
  const res=await fetch('/api/email/status');
  const d=await res.json();
  const log=document.getElementById('gen-log');
  (d.new_logs||[]).forEach(l=>{
    const div=document.createElement('div');
    div.className='l '+logClass(l); div.textContent=l;
    if(log){log.appendChild(div);log.scrollTop=log.scrollHeight;}
  });
  if(d.total>0){
    const pct=Math.round(d.done/d.total*100);
    document.getElementById('prog-fill').style.width=pct+'%';
    document.getElementById('prog-txt').textContent=`${d.done}/${d.total} emails`;
  }
  if(d.status!=='running'){
    clearInterval(genPoll);
    document.getElementById('btn-gen').disabled=false;
    document.getElementById('btn-gen').textContent='🤖 Gen AI Email';
    if(d.status==='done'&&d.leads){
      emailLeads=d.leads;
      document.getElementById('cnt-email').textContent=emailLeads.filter(l=>l.email_body).length;
      renderEmails(emailLeads);
      document.getElementById('btn-export').disabled=false;
    }
  }
}
function renderEmails(leads){
  const w=leads.filter(l=>l.email_body);
  let html=`<div style="font-size:.8rem;color:var(--muted);margin-bottom:14px">${w.length}/${leads.length} leads đã có email</div>`;
  w.forEach(l=>{
    const site=(()=>{try{return new URL(l.website).hostname.replace('www.','')}catch{return l.website}})();
    html+=`<div class="email-card">
      <div class="ec-site">${site}</div>
      <div class="ec-subj">📧 ${l.email_subject||'(no subject)'}</div>
      <div class="ec-body">${(l.email_body||'').replace(/</g,'&lt;')}</div>
      ${l.pain_point?`<div class="ec-pain">💡 ${l.pain_point}</div>`:''}
      <div class="ec-meta">${l.field||''} · ${l.best_email||''} · ${l.ai_language||''}</div>
    </div>`;
  });
  document.getElementById('email-panel').innerHTML=html||'<div class="empty"><div class="ei">📭</div>Không có email nào được gen.</div>';
}

// ── EXPORT ────────────────────────────────────────────────────────────
async function exportExcel(){
  const leads=emailLeads.length?emailLeads:filteredLeads.length?filteredLeads:crawlLeads;
  if(!leads.length){alert('Không có dữ liệu.');return;}
  const btn=document.getElementById('btn-export');
  btn.textContent='⏳ Xuất...'; btn.disabled=true;
  const res=await fetch('/api/export',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({leads})});
  const d=await res.json();
  if(d.ok) window.location.href='/api/download/'+d.filename;
  btn.textContent='⬇ Export Excel'; btn.disabled=false;
}

// ── SESSIONS ─────────────────────────────────────────────────────────
async function loadSessions(){
  const res=await fetch('/api/sessions');
  const arr=await res.json();
  const el=document.getElementById('sessions');
  if(!arr.length){el.textContent='—';return;}
  el.innerHTML=arr.slice(0,5).map(s=>`
    <div style="padding:5px 0;border-bottom:1px solid var(--border)">
      <span style="color:var(--text)">${s.industry}</span>
      <span style="color:var(--muted);float:right">${s.leads} leads</span>
      <div style="font-size:.67rem;color:#374151">${s.session}</div>
    </div>`).join('');
}

// ── INIT ──────────────────────────────────────────────────────────────
renderIndustries();
selIndustry('hospitality');
loadSessions();
goTab('crawl');
</script>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════
#  FLASK ROUTES
# ══════════════════════════════════════════════════════════════════════
@app.route("/")
def index():
    meta = {
        k: {**v, "cats": len(YP_CATEGORIES.get(k, []))}
        for k, v in INDUSTRY_META.items()
    }
    return render_template_string(HTML, industries=meta)


@app.route("/api/crawl/start", methods=["POST"])
def api_crawl_start():
    global _crawl
    with _lock:
        if _crawl.get("status") == "running":
            return jsonify({"ok": False, "msg": "Already running"})
        _crawl = {
            "status": "running",
            "logs": [],
            "_log_cursor": 0,
            "leads": [],
            "session": "",
            "total_urls": 0,
            "done_urls": 0,
            "phase": "Khởi động...",
        }
    threading.Thread(target=run_crawl_job, args=(request.json,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/crawl/stop", methods=["POST"])
def api_crawl_stop():
    with _lock:
        _crawl["status"] = "stopping"
    return jsonify({"ok": True})


@app.route("/api/crawl/status")
def api_crawl_status():
    with _lock:
        cursor = _crawl.get("_log_cursor", 0)
        all_logs = _crawl.get("logs", [])
        new_logs = all_logs[cursor:]
        _crawl["_log_cursor"] = len(all_logs)
        return jsonify(
            {
                "status": _crawl.get("status", "idle"),
                "new_logs": new_logs,
                "leads": _crawl.get("leads", []),
                "total_urls": _crawl.get("total_urls", 0),
                "done_urls": _crawl.get("done_urls", 0),
                "phase": _crawl.get("phase", ""),
            }
        )


@app.route("/api/filter/upload", methods=["POST"])
def api_filter_upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "no file"})
    try:
        df = (
            pd.read_csv(f, dtype=str).fillna("")
            if f.filename.endswith(".csv")
            else pd.read_excel(f, dtype=str).fillna("")
        )
        with _lock:
            _filter_state["df_raw"] = df
        return jsonify({"ok": True, "rows": len(df)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/filter/load-crawl", methods=["POST"])
def api_filter_load_crawl():
    leads = request.json.get("leads", [])
    if leads:
        with _lock:
            _filter_state["df_raw"] = pd.DataFrame(leads).fillna("")
    return jsonify({"ok": True})


@app.route("/api/filter/run", methods=["POST"])
def api_filter_run():
    with _lock:
        df = _filter_state.get("df_raw")
    if df is None:
        return jsonify(
            {"ok": False, "error": "Chưa có dữ liệu. Upload file hoặc crawl trước."}
        )
    cfg = request.json or {}
    filter_cfg = {
        "min_score": cfg.get("min_score", 15),
        "grades": [g.upper() for g in cfg.get("grades", ["A", "B", "C"])],
        "require_email": cfg.get("require_email", True),
        "require_phone": cfg.get("require_phone", False),
        "industries": cfg.get("industries", []),
    }
    try:
        df_pass, df_rej = apply_filter(df.copy(), filter_cfg)
        with _lock:
            _filter_state["df_filtered"] = df_pass
            _filter_state["df_rejected"] = df_rej
        n = len(df_pass)
        calls = max(1, n // 5 + 1)
        cost = (calls * 500 * 0.15 + calls * 1000 * 0.60) / 1_000_000
        return jsonify(
            {
                "ok": True,
                "passed": (
                    df_pass.to_dict(orient="records") if not df_pass.empty else []
                ),
                "rejected": (
                    df_rej.to_dict(orient="records") if not df_rej.empty else []
                ),
                "cost_estimate": f"~${cost:.4f} USD ({calls} API calls)",
            }
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/email/gen", methods=["POST"])
def api_email_gen():
    global _gen
    with _lock:
        if _gen.get("status") == "running":
            return jsonify({"ok": False, "error": "Already running"})
        _gen = {
            "status": "running",
            "logs": ["🤖 Khởi động..."],
            "_log_cursor": 0,
            "progress": 0,
            "total": 0,
            "leads": [],
        }
    leads = request.json.get("leads", [])

    def run():
        global _gen

        def glog(m):
            with _lock:
                _gen["logs"].append(m)

        def progress(done, total):
            with _lock:
                _gen["progress"] = done
                _gen["total"] = total
            glog(f"  ✉️  {done}/{total} emails gen xong")

        try:
            from email_generator import gen_emails_for_leads

            glog(f"📦 {len(leads)} leads | GPT-4o-mini")
            result = gen_emails_for_leads(leads, progress_cb=progress)
            with_em = sum(1 for r in result if r.get("email_body"))
            glog(f"✅ Hoàn thành | {with_em}/{len(result)} emails đã gen")
            with _lock:
                _gen["leads"] = result
                _gen["status"] = "done"
        except Exception as e:
            with _lock:
                _gen["logs"].append(f"❌ Lỗi: {e}")
                _gen["status"] = "error"

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/email/status")
def api_email_status():
    with _lock:
        cursor = _gen.get("_log_cursor", 0)
        all_logs = _gen.get("logs", [])
        new_logs = all_logs[cursor:]
        _gen["_log_cursor"] = len(all_logs)
        return jsonify(
            {
                "status": _gen.get("status", "idle"),
                "done": _gen.get("progress", 0),
                "total": _gen.get("total", 0),
                "new_logs": new_logs,
                "leads": _gen.get("leads", []),
            }
        )


@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.json.get("leads", [])
    if not data:
        return jsonify({"ok": False, "error": "no data"})
    df_pass = pd.DataFrame(data)
    with _lock:
        df_rej = _filter_state.get("df_rejected", pd.DataFrame())
    filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        _write_excel(
            df_pass, df_rej if df_rej is not None else pd.DataFrame(), filename
        )
        return jsonify({"ok": True, "filename": filename})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/download/<filename>")
def api_download(filename):
    safe = os.path.basename(filename)
    if os.path.exists(safe) and safe.endswith((".xlsx", ".csv")):
        return send_file(safe, as_attachment=True, download_name=safe)
    return "File không tồn tại.", 404


@app.route("/api/sessions")
def api_sessions():
    try:
        return jsonify(json.load(open("sessions.json", "r", encoding="utf-8")))
    except:
        return jsonify([])


if __name__ == "__main__":
    print("\n" + "=" * 52)
    print("  🎯 Lead Intelligence Tool v5")
    print("  Nguồn: YellowPages.vn → website thật")
    print("  http://localhost:5000")
    print("=" * 52 + "\n")
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)
